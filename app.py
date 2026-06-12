"""
Fertilizer Inventory Management System (FIMS) - Redesigned
Role-based application with specific dashboards
"""

import os
import random
import requests as http_requests
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from urllib.parse import unquote
from datetime import datetime, timedelta
from database_pg import Database

# Get the directory where app.py is located
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, 
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.secret_key = os.environ.get('SECRET_KEY', 'fims-secret-key-change-in-production')

# Initialize database
db = Database()
db.initialize_database()

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ── Email / Notification helpers ────────────────────────────────────────────
NOTIFY_EMAIL       = os.environ.get('NOTIFY_EMAIL', 'dipeshasrani9@gmail.com')
FORMSPREE_FORM_ID  = os.environ.get('FORMSPREE_FORM_ID', '')

def send_formspree(subject: str, message: str) -> bool:
    """Send an email via Formspree API. Returns True on success."""
    if not FORMSPREE_FORM_ID:
        app.logger.warning("FORMSPREE_FORM_ID not set — email not sent")
        return False
    try:
        resp = http_requests.post(
            f'https://formspree.io/f/{FORMSPREE_FORM_ID}',
            json={'email': NOTIFY_EMAIL, '_subject': subject, 'message': message},
            headers={'Accept': 'application/json'},
            timeout=8
        )
        return resp.status_code == 200
    except Exception as e:
        app.logger.error(f"Formspree send failed: {e}")
        return False

# ── User class for Flask-Login
class User(UserMixin):
    def __init__(self, user_id, username, role):
        self.id = user_id
        self.username = username
        # Map lowercase PostgreSQL enum values to capitalized app values
        role_map = {'admin': 'Admin', 'rakepoint': 'RakePoint', 'warehouse': 'Warehouse', 'accountant': 'Accountant'}
        self.role = role_map.get(role, role)

@login_manager.user_loader
def load_user(user_id):
    user_data = db.get_user_by_id(user_id)
    if user_data:
        return User(user_data[0], user_data[1], user_data[3])
    return None

# Context processor to make datetime available in all templates
@app.context_processor
def inject_datetime():
    return {'datetime': datetime}

# Custom Jinja2 filter for date formatting (dd-mm-yy)
@app.template_filter('format_date')
def format_date_filter(value):
    """Convert date to dd-mm-yy format"""
    if value is None:
        return ''
    if isinstance(value, str):
        # Try to parse common date formats
        for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y', '%d/%m/%Y']:
            try:
                value = datetime.strptime(value.split()[0] if ' ' in value else value, fmt)
                break
            except ValueError:
                continue
        else:
            return value[:10] if len(value) >= 10 else value  # Return as-is if can't parse
    if isinstance(value, datetime):
        return value.strftime('%d-%m-%y')
    return str(value)

# ========== Authentication Routes ==========

@app.route('/')
def index():
    if current_user.is_authenticated:
        # Redirect to role-specific dashboard
        if current_user.role == 'Admin':
            return redirect(url_for('admin_dashboard'))
        elif current_user.role == 'RakePoint':
            return redirect(url_for('rakepoint_dashboard'))
        elif current_user.role == 'Warehouse':
            return redirect(url_for('warehouse_dashboard'))
        elif current_user.role == 'Accountant':
            return redirect(url_for('accountant_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user_data = db.authenticate_user(username, password)
        if user_data:
            user = User(user_data[0], user_data[1], user_data[3])
            login_user(user)
            flash('Login successful!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))

# ========== ADMIN Dashboard & Routes ==========

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Use optimized single-query method for stats
    stats = db.get_dashboard_stats_optimized()
    
    # Get total shortage from closed rakes
    total_shortage = db.get_total_shortage()
    
    # Use optimized method - single query for all rakes with balances (limit 5)
    recent_rakes = db.get_rakes_with_balances(limit=5)
    
    # Get day-wise warehouse stock for last 7 days
    daywise_stock = db.get_daywise_warehouse_stock(days=7)
    
    return render_template('admin/dashboard.html', stats=stats, recent_rakes=recent_rakes, total_shortage=total_shortage, daywise_stock=daywise_stock)

@app.route('/admin/add-rake', methods=['GET', 'POST'])
@login_required
def admin_add_rake():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        rake_code = request.form.get('rake_code')
        company_name = request.form.get('company_name')
        company_code = request.form.get('company_code')
        date = request.form.get('date')
        rr_quantity = float(request.form.get('rr_quantity'))
        rake_point_name = request.form.get('rake_point_name')
        builty_head = request.form.get('builty_head')
        
        # Get multiple products
        product_ids = request.form.getlist('product_ids[]')
        product_names = request.form.getlist('product_names[]')
        product_codes = request.form.getlist('product_codes[]')
        product_quantities = request.form.getlist('product_quantities[]')
        
        # Build products list
        products = []
        for i in range(len(product_ids)):
            if product_ids[i] and product_names[i]:
                products.append({
                    'product_id': product_ids[i] if product_ids[i] else None,
                    'product_name': product_names[i],
                    'product_code': product_codes[i] if i < len(product_codes) else '',
                    'quantity_mt': float(product_quantities[i]) if i < len(product_quantities) and product_quantities[i] else 0
                })
        
        # Use first product as primary for backward compatibility
        product_name = products[0]['product_name'] if products else ''
        product_code = products[0]['product_code'] if products else ''
        
        rake_id = db.add_rake(rake_code, company_name, company_code, date, rr_quantity,
                             product_name, product_code, rake_point_name, builty_head, products)
        
        if rake_id:
            flash(f'Rake {rake_code} added successfully with {len(products)} product(s)!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Error adding rake. Rake code may already exist.', 'error')
    
    products = db.get_all_products()
    companies = db.get_all_companies()
    return render_template('admin/add_rake.html', products=products, companies=companies)

@app.route('/admin/edit-rake/<path:rake_code>', methods=['GET', 'POST'])
@login_required
def admin_edit_rake(rake_code):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    rake_code = unquote(rake_code)
    
    if request.method == 'POST':
        company_name = request.form.get('company_name')
        company_code = request.form.get('company_code', '')
        date = request.form.get('date')
        rr_quantity = float(request.form.get('rr_quantity'))
        rake_point_name = request.form.get('rake_point_name')
        builty_head = request.form.get('builty_head', '')
        
        product_ids = request.form.getlist('product_ids[]')
        product_names = request.form.getlist('product_names[]')
        product_codes = request.form.getlist('product_codes[]')
        product_quantities = request.form.getlist('product_quantities[]')
        
        products = []
        for i in range(len(product_ids)):
            if product_ids[i] and product_names[i]:
                products.append({
                    'product_id': product_ids[i] if product_ids[i] else None,
                    'product_name': product_names[i],
                    'product_code': product_codes[i] if i < len(product_codes) else '',
                    'quantity_mt': float(product_quantities[i]) if i < len(product_quantities) and product_quantities[i] else 0
                })
        
        product_name = products[0]['product_name'] if products else ''
        product_code = products[0]['product_code'] if products else ''
        
        success = db.update_rake(rake_code, company_name, company_code, date, rr_quantity,
                                 product_name, product_code, rake_point_name, builty_head, products)
        
        if success:
            db.invalidate_cache()
            flash(f'Rake {rake_code} updated successfully!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Error updating rake. Please try again.', 'error')
    
    rake = db.get_rake_by_code(rake_code)
    if not rake:
        flash('Rake not found', 'error')
        return redirect(url_for('admin_dashboard'))
    
    existing_products = db.get_rake_products(rake_code)
    products = db.get_all_products()
    companies = db.get_all_companies()
    return render_template('admin/edit_rake.html', rake=rake, existing_products=existing_products,
                           products=products, companies=companies)

@app.route('/admin/close-rake/<path:rake_code>', methods=['POST'])
@login_required
def admin_close_rake(rake_code):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    success, result = db.close_rake(rake_code)
    
    if success:
        flash(f'Rake {rake_code} closed successfully! Shortage: {result:.2f} MT', 'success')
    else:
        flash(f'Error closing rake: {result}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/reopen-rake/<path:rake_code>', methods=['POST'])
@login_required
def admin_reopen_rake(rake_code):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    success = db.reopen_rake(rake_code)
    
    if success:
        flash(f'Rake {rake_code} reopened successfully!', 'success')
    else:
        flash(f'Error reopening rake', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/summary')
@login_required
def admin_summary():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    view_type = request.args.get('view', 'rakes')  # rakes, total, per_account, do
    
    # Get total shortage from closed rakes
    total_shortage = db.get_total_shortage()
    
    # Use optimized method - single query instead of N+1
    rakes_with_balance = db.get_rakes_with_balances()
    
    # Get total summary (all accounts with dispatch quantities) - use loading_slips
    total_account_summary = db.execute_custom_query('''
        SELECT a.account_name, a.account_type, 
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN accounts a ON ls.account_id = a.id
        GROUP BY a.id, a.account_name, a.account_type
        ORDER BY total_quantity DESC
    ''')
    
    # Get CGMF dispatch summary - use loading_slips
    cgmf_summary = db.execute_custom_query('''
        SELECT c.society_name, c.district, c.destination,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN cgmf c ON ls.cgmf_id = c.id
        GROUP BY c.id, c.society_name, c.district
        ORDER BY total_quantity DESC
    ''')
    
    # Get warehouse dispatch summary - use loading_slips
    warehouse_summary = db.execute_custom_query('''
        SELECT w.warehouse_name, w.location,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN warehouses w ON ls.warehouse_id = w.id
        GROUP BY w.id, w.warehouse_name
        ORDER BY total_quantity DESC
    ''')

    # ── DO Summary (carry-forward per account per rake) ──────────────
    do_summary_rows = []
    do_filter_accounts = []
    do_filter_rakes = []
    do_filter_products = []

    if view_type == 'do':
        from collections import OrderedDict
        raw_dos = db.execute_custom_query('''
            SELECT do_t.account_id::text, a.account_name, a.account_type,
                   do_t.rake_code, r.date AS rake_date,
                   do_t.product_name, do_t.quantity_bags, do_t.quantity_mt,
                   do_t.id::text
            FROM dispatch_orders do_t
            LEFT JOIN accounts a ON do_t.account_id = a.id
            JOIN rakes r ON do_t.rake_code = r.rake_code
            ORDER BY a.account_name, do_t.product_name, r.date, do_t.rake_code
        ''') or []

        dispatched_raw = db.execute_custom_query('''
            SELECT b.account_id::text, b.rake_code,
                   SUM(b.quantity_mt)::numeric    AS dispatched_mt,
                   SUM(b.number_of_bags)::integer AS dispatched_bags
            FROM builty b
            WHERE b.account_id IS NOT NULL AND b.rake_code IS NOT NULL
            GROUP BY b.account_id, b.rake_code
        ''') or []
        dispatched_map = {
            (r[0], r[1]): (float(r[2] or 0), int(r[3] or 0))
            for r in dispatched_raw
        }

        # Group DOs by (account_id, product_name) in insertion order
        groups = OrderedDict()
        for do in raw_dos:
            key = (do[0], do[5])  # (account_id, product_name)
            groups.setdefault(key, []).append(do)

        for (account_id, product), do_list in groups.items():
            carry_forward = 0.0
            for do in do_list:
                rake_code    = do[3]
                account_name = do[1]
                account_type = do[2]
                do_mt        = float(do[7] or 0)
                do_bags      = int(do[6] or 0)
                rake_date    = do[4]
                do_id        = do[8]

                disp_mt, disp_bags = dispatched_map.get((account_id, rake_code), (0.0, 0))
                total_expected = carry_forward + do_mt
                difference     = total_expected - disp_mt

                do_summary_rows.append({
                    'do_id':          do_id,
                    'account_name':   account_name,
                    'account_type':   account_type,
                    'product':        product,
                    'rake_code':      rake_code,
                    'rake_date':      rake_date,
                    'do_qty_mt':      round(do_mt, 2),
                    'do_qty_bags':    do_bags,
                    'carry_forward':  round(carry_forward, 2),
                    'total_expected': round(total_expected, 2),
                    'dispatched_mt':  round(disp_mt, 2),
                    'dispatched_bags':disp_bags,
                    'difference':     round(difference, 2),
                })

                carry_forward = max(0.0, difference)

        do_filter_accounts = sorted({r['account_name'] for r in do_summary_rows})
        do_filter_rakes    = sorted({r['rake_code']    for r in do_summary_rows})
        do_filter_products = sorted({r['product']      for r in do_summary_rows})

    return render_template('admin/summary.html',
                         rakes=rakes_with_balance,
                         view_type=view_type,
                         total_account_summary=total_account_summary,
                         cgmf_summary=cgmf_summary,
                         warehouse_summary=warehouse_summary,
                         total_shortage=total_shortage or 0,
                         do_summary_rows=do_summary_rows,
                         do_filter_accounts=do_filter_accounts,
                         do_filter_rakes=do_filter_rakes,
                         do_filter_products=do_filter_products)

@app.route('/admin/do', methods=['GET', 'POST'])
@login_required
def admin_do():
    """Admin: create / list Dispatch Orders"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action', 'create')

        if action == 'delete':
            do_id = request.form.get('do_id')
            if do_id and db.delete_dispatch_order(do_id):
                flash('Dispatch Order deleted.', 'success')
            else:
                flash('Failed to delete Dispatch Order.', 'error')
            return redirect(url_for('admin_do'))

        # Create
        dest_type    = request.form.get('dest_type', 'account')
        account_id   = request.form.get('account_id', '').strip() or None
        cgmf_id      = request.form.get('cgmf_id', '').strip() or None
        rake_code    = request.form.get('rake_code', '').strip()
        product_name = request.form.get('product_name', '').strip()
        notes        = request.form.get('notes', '').strip()
        try:
            quantity_bags = int(request.form.get('quantity_bags', 0) or 0)
            quantity_mt   = float(request.form.get('quantity_mt', 0) or 0)
        except (ValueError, TypeError):
            flash('Invalid quantity values.', 'error')
            return redirect(url_for('admin_do'))

        if dest_type == 'cgmf':
            account_id = None
        else:
            cgmf_id = None

        if not (rake_code and product_name and quantity_mt > 0 and (account_id or cgmf_id)):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('admin_do'))

        result = db.create_dispatch_order(
            account_id=account_id,
            rake_code=rake_code,
            product_name=product_name,
            quantity_bags=quantity_bags,
            quantity_mt=quantity_mt,
            notes=notes,
            created_by=current_user.username,
            cgmf_id=cgmf_id
        )
        if result:
            flash('Dispatch Order created successfully.', 'success')
        else:
            flash('Failed to create Dispatch Order.', 'error')
        return redirect(url_for('admin_do'))

    accounts  = db.get_all_accounts()
    rakes     = db.get_rakes_with_balances()
    cgmf_list = db.get_all_cgmf()
    dos       = db.get_all_dispatch_orders() or []
    return render_template('admin/do_entry.html', accounts=accounts, rakes=rakes, dos=dos, cgmf_list=cgmf_list)


@app.route('/admin/rake-details/<path:rake_code>')
@login_required
def admin_rake_details(rake_code):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    # Get rake basic info
    rake_info = db.execute_custom_query('''
        SELECT r.rake_code, r.company_name, r.product_name, r.date, 
               r.rr_quantity, r.rake_point_name
        FROM rakes r
        WHERE r.rake_code = ?
    ''', (rake_code,))
    
    if not rake_info:
        flash('Rake not found', 'error')
        return redirect(url_for('admin_summary'))
    
    # Get rake products (for multi-product support)
    rake_products = db.get_rake_products(rake_code)
    
    # Get stock dispatched to accounts FROM LOADING SLIPS (the actual dispatch record)
    account_dispatches = db.execute_custom_query('''
        SELECT a.account_name, a.account_type,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(ls.id) as slip_count,
               STRING_AGG(COALESCE('Builty#' || b.builty_number, 'Slip#' || ls.slip_number::text) || ' (' || ls.quantity_mt::text || ' MT)', ', ') as slip_details,
               a.address
        FROM loading_slips ls
        JOIN accounts a ON ls.account_id = a.id
        LEFT JOIN builty b ON ls.builty_id = b.id
        WHERE ls.rake_code = ?
        GROUP BY a.id, a.account_name, a.account_type, a.address
        ORDER BY total_quantity DESC
    ''', (rake_code,))
    
    # Get stock sent to warehouses FROM LOADING SLIPS
    warehouse_dispatches = db.execute_custom_query('''
        SELECT w.warehouse_name, w.location,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(ls.id) as slip_count,
               STRING_AGG(COALESCE('Builty#' || b.builty_number, 'Slip#' || ls.slip_number::text) || ' (' || ls.quantity_mt::text || ' MT)', ', ') as slip_details
        FROM loading_slips ls
        JOIN warehouses w ON ls.warehouse_id = w.id
        LEFT JOIN builty b ON ls.builty_id = b.id
        WHERE ls.rake_code = ?
        GROUP BY w.id, w.warehouse_name, w.location
        ORDER BY total_quantity DESC
    ''', (rake_code,))
    
    # Get CGMF dispatches FROM LOADING SLIPS
    cgmf_dispatches = db.execute_custom_query('''
        SELECT c.society_name, c.district, c.destination,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(ls.id) as slip_count,
               STRING_AGG(COALESCE('Builty#' || b.builty_number, 'Slip#' || ls.slip_number::text) || ' (' || ls.quantity_mt::text || ' MT)', ', ') as slip_details
        FROM loading_slips ls
        JOIN cgmf c ON ls.cgmf_id = c.id
        LEFT JOIN builty b ON ls.builty_id = b.id
        WHERE ls.rake_code = ?
        GROUP BY c.id, c.society_name, c.district
        ORDER BY total_quantity DESC
    ''', (rake_code,))
    
    # Calculate totals
    total_to_accounts = sum(row[2] for row in account_dispatches) if account_dispatches else 0
    total_to_warehouses = sum(row[2] for row in warehouse_dispatches) if warehouse_dispatches else 0
    total_to_cgmf = sum(row[3] for row in cgmf_dispatches) if cgmf_dispatches else 0

    # All individual builties for this rake (for the builty detail report)
    rake_builties = db.execute_custom_query('''
        SELECT b.builty_number, b.date, b.lr_number,
               COALESCE(a.account_name, w.warehouse_name, c.society_name, 'N/A') AS account,
               b.sub_head, t.truck_number, b.quantity_mt, b.number_of_bags,
               CASE
                   WHEN a.id IS NOT NULL THEN a.account_type::TEXT
                   WHEN w.id IS NOT NULL THEN 'Warehouse'
                   WHEN c.id IS NOT NULL THEN 'CGMF'
                   ELSE 'Unknown'
               END AS dispatch_type,
               COALESCE(b.total_freight, 0) AS total_freight
        FROM builty b
        LEFT JOIN accounts   a ON b.account_id  = a.id
        LEFT JOIN warehouses w ON b.warehouse_id = w.id
        LEFT JOIN cgmf       c ON b.cgmf_id     = c.id
        LEFT JOIN trucks     t ON b.truck_id     = t.id
        WHERE b.rake_code = %s
        ORDER BY b.date, b.builty_number
    ''', (rake_code,))

    return render_template('admin/rake_details.html',
                         rake_info=rake_info[0],
                         rake_products=rake_products,
                         account_dispatches=account_dispatches,
                         warehouse_dispatches=warehouse_dispatches,
                         cgmf_dispatches=cgmf_dispatches,
                         total_to_accounts=total_to_accounts,
                         total_to_warehouses=total_to_warehouses,
                         total_to_cgmf=total_to_cgmf,
                         rake_builties=rake_builties or [],
                         notify_email=NOTIFY_EMAIL)

@app.route('/admin/send-delete-rake-otp/<path:rake_code>', methods=['POST'])
@login_required
def admin_send_delete_rake_otp(rake_code):
    """Generate a 6-digit OTP, store in session, and email it via Formspree."""
    if current_user.role != 'Admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    rake_code = unquote(rake_code)
    otp = str(random.randint(100000, 999999))
    session['delete_rake_otp']       = otp
    session['delete_rake_otp_code']  = rake_code
    session['delete_rake_otp_expiry'] = (datetime.utcnow() + timedelta(minutes=10)).isoformat()

    sent = send_formspree(
        subject=f'[FIMS] OTP to Delete Rake {rake_code}',
        message=(
            f'An admin has requested to delete rake {rake_code} and ALL its associated '
            f'loading slips, builties and e-bills.\n\n'
            f'Your one-time password is: {otp}\n\n'
            f'This OTP expires in 10 minutes. If you did not request this, please contact the administrator immediately.'
        )
    )
    if sent:
        return jsonify({'success': True, 'message': f'OTP sent to {NOTIFY_EMAIL}'})
    else:
        return jsonify({'success': False, 'message': 'Failed to send OTP email. Check FORMSPREE_FORM_ID env var.'})

@app.route('/admin/delete-rake/<path:rake_code>', methods=['POST'])
@login_required
def admin_delete_rake(rake_code):
    """Delete a rake after OTP verification."""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    rake_code = unquote(rake_code)

    entered_otp   = request.form.get('otp', '').strip()
    stored_otp    = session.get('delete_rake_otp')
    stored_code   = session.get('delete_rake_otp_code')
    expiry_str    = session.get('delete_rake_otp_expiry')

    if not stored_otp or stored_code != rake_code:
        flash('No OTP found for this rake. Please request a new OTP.', 'error')
        return redirect(url_for('admin_rake_details', rake_code=rake_code))

    if expiry_str and datetime.utcnow() > datetime.fromisoformat(expiry_str):
        session.pop('delete_rake_otp', None)
        flash('OTP has expired. Please request a new one.', 'error')
        return redirect(url_for('admin_rake_details', rake_code=rake_code))

    if entered_otp != stored_otp:
        flash('Invalid OTP. Please try again.', 'error')
        return redirect(url_for('admin_rake_details', rake_code=rake_code))

    # OTP valid — clear session and delete
    session.pop('delete_rake_otp', None)
    session.pop('delete_rake_otp_code', None)
    session.pop('delete_rake_otp_expiry', None)

    success, message = db.delete_rake(rake_code)
    if success:
        send_formspree(
            subject=f'[FIMS] Rake {rake_code} Deleted',
            message=f'Admin {current_user.username} deleted rake {rake_code} and all its associated records.'
        )
        flash(f'Rake {rake_code} and all associated data deleted successfully.', 'success')
        return redirect(url_for('admin_summary'))
    else:
        flash(f'Error deleting rake: {message}', 'error')
        return redirect(url_for('admin_rake_details', rake_code=rake_code))

@app.route('/admin/download-rake-summary-excel')
@login_required
def admin_download_rake_summary_excel():
    """Download all rakes summary as Excel file"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    # Get all rakes with balance
    all_rakes = db.get_all_rakes()
    total_shortage = db.get_total_shortage() or 0
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rake Summary"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    
    # Title
    ws['A1'] = "RAKE SUMMARY REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:J1')
    
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True)
    
    # Headers
    headers = ['Rake Code', 'Company', 'Product', 'Date', 'RR Quantity (MT)', 'Dispatched (MT)', 'Remaining (MT)', 'Shortage (MT)', 'Status', 'Rake Point']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    total_rr = 0
    total_dispatched = 0
    total_remaining = 0
    
    for rake in all_rakes:
        rake_code = rake[1]
        balance = db.get_rake_balance(rake_code)
        is_closed = rake[10] if len(rake) > 10 else 0
        shortage = rake[12] if len(rake) > 12 else 0
        
        dispatched = balance['dispatched'] if balance else 0
        remaining = balance['remaining'] if balance else rake[5]
        status = 'Closed' if is_closed else ('Active' if remaining > 0 else 'Completed')
        
        total_rr += rake[5]
        total_dispatched += dispatched
        total_remaining += remaining
        
        ws.cell(row=row, column=1, value=rake[1]).border = border  # rake_code
        ws.cell(row=row, column=2, value=rake[2]).border = border  # company_name
        ws.cell(row=row, column=3, value=rake[6]).border = border  # product_name
        ws.cell(row=row, column=4, value=rake[4]).border = border  # date
        ws.cell(row=row, column=5, value=round(rake[5], 2)).border = border  # rr_quantity
        ws.cell(row=row, column=6, value=round(dispatched, 2)).border = border
        ws.cell(row=row, column=7, value=round(remaining, 2)).border = border
        ws.cell(row=row, column=8, value=round(shortage, 2) if is_closed else '-').border = border
        ws.cell(row=row, column=9, value=status).border = border
        ws.cell(row=row, column=10, value=rake[8]).border = border  # rake_point_name
        row += 1
    
    # Totals row
    for col in range(1, 11):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=col).border = border
    
    ws.cell(row=row, column=1, value="GRAND TOTAL")
    ws.cell(row=row, column=5, value=round(total_rr, 2))
    ws.cell(row=row, column=6, value=round(total_dispatched, 2))
    ws.cell(row=row, column=7, value=round(total_remaining, 2))
    ws.cell(row=row, column=8, value=round(total_shortage, 2))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'rake_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/download-rake-details-excel/<path:rake_code>')
@login_required
def admin_download_rake_details_excel(rake_code):
    """Download individual rake details as Excel file"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    # Get rake basic info
    rake_info = db.execute_custom_query('''
        SELECT r.rake_code, r.company_name, r.product_name, r.date, 
               r.rr_quantity, r.rake_point_name
        FROM rakes r
        WHERE r.rake_code = ?
    ''', (rake_code,))
    
    if not rake_info:
        flash('Rake not found', 'error')
        return redirect(url_for('admin_summary'))
    
    rake_info = rake_info[0]
    
    # Get dispatches — summary per account/warehouse/CGMF with all builty numbers
    account_dispatches = db.execute_custom_query('''
        SELECT a.account_name, a.account_type,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(ls.id) as slip_count,
               COUNT(b.id) as builty_count,
               STRING_AGG(b.builty_number, ', ' ORDER BY b.builty_number) as builty_numbers,
               a.address
        FROM loading_slips ls
        JOIN accounts a ON ls.account_id = a.id
        LEFT JOIN builty b ON ls.builty_id = b.id
        WHERE ls.rake_code = ?
        GROUP BY a.id, a.account_name, a.account_type, a.address
        ORDER BY total_quantity DESC
    ''', (rake_code,))

    warehouse_dispatches = db.execute_custom_query('''
        SELECT w.warehouse_name, w.location,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(ls.id) as slip_count,
               COUNT(b.id) as builty_count,
               STRING_AGG(b.builty_number, ', ' ORDER BY b.builty_number) as builty_numbers
        FROM loading_slips ls
        JOIN warehouses w ON ls.warehouse_id = w.id
        LEFT JOIN builty b ON ls.builty_id = b.id
        WHERE ls.rake_code = ?
        GROUP BY w.id, w.warehouse_name, w.location
        ORDER BY total_quantity DESC
    ''', (rake_code,))

    cgmf_dispatches = db.execute_custom_query('''
        SELECT c.society_name, c.district, c.destination,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(ls.id) as slip_count,
               COUNT(b.id) as builty_count,
               STRING_AGG(b.builty_number, ', ' ORDER BY b.builty_number) as builty_numbers
        FROM loading_slips ls
        JOIN cgmf c ON ls.cgmf_id = c.id
        LEFT JOIN builty b ON ls.builty_id = b.id
        WHERE ls.rake_code = ?
        GROUP BY c.id, c.society_name, c.district
        ORDER BY total_quantity DESC
    ''', (rake_code,))

    # All individual builties for this rake (detail sheet)
    all_builties = db.execute_custom_query('''
        SELECT b.builty_number, b.date, b.lr_number,
               COALESCE(a.account_name, w.warehouse_name, c.society_name, 'N/A') AS recipient,
               b.sub_head, t.truck_number, b.number_of_bags, b.quantity_mt,
               b.loading_point, b.unloading_point, b.total_freight
        FROM builty b
        LEFT JOIN accounts a   ON b.account_id   = a.id
        LEFT JOIN warehouses w ON b.warehouse_id  = w.id
        LEFT JOIN cgmf c       ON b.cgmf_id       = c.id
        LEFT JOIN trucks t     ON b.truck_id       = t.id
        WHERE b.rake_code = ?
        ORDER BY b.date, b.builty_number
    ''', (rake_code,))
    
    # Calculate totals
    total_to_accounts = sum(row[2] for row in account_dispatches) if account_dispatches else 0
    total_to_warehouses = sum(row[2] for row in warehouse_dispatches) if warehouse_dispatches else 0
    total_to_cgmf = sum(row[3] for row in cgmf_dispatches) if cgmf_dispatches else 0
    total_dispatched = total_to_accounts + total_to_warehouses + total_to_cgmf
    remaining = rake_info[4] - total_dispatched
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rake {rake_code}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    section_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    total_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"RAKE DETAILS - {rake_code}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    # Rake Info
    row = 3
    ws.cell(row=row, column=1, value="Company:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=rake_info[1])
    ws.cell(row=row, column=3, value="Product:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=rake_info[2])
    row += 1
    ws.cell(row=row, column=1, value="Date:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=rake_info[3])
    ws.cell(row=row, column=3, value="RR Quantity:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"{rake_info[4]} MT")
    row += 1
    ws.cell(row=row, column=1, value="Rake Point:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=rake_info[5])
    row += 2
    
    # Accounts Section
    ws.cell(row=row, column=1, value="DISPATCHED TO ACCOUNTS").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    
    acc_headers = ['Type', 'Account Name', 'Location', 'Quantity (MT)', 'Loading Slips', 'Builties', 'Builty Numbers']
    for col, header in enumerate(acc_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    for acc in account_dispatches:
        ws.cell(row=row, column=1, value=acc[1]).border = border  # type
        ws.cell(row=row, column=2, value=acc[0]).border = border  # account_name
        ws.cell(row=row, column=3, value=acc[6] or '').border = border  # address
        ws.cell(row=row, column=4, value=round(acc[2], 2)).border = border
        ws.cell(row=row, column=5, value=acc[3]).border = border
        ws.cell(row=row, column=6, value=acc[4]).border = border
        ws.cell(row=row, column=7, value=acc[5] or '-').border = border
        row += 1
    
    ws.cell(row=row, column=1, value="Total to Accounts:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_to_accounts, 2)).font = Font(bold=True)
    row += 2
    
    # Warehouses Section
    ws.cell(row=row, column=1, value="DISPATCHED TO WAREHOUSES").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    
    wh_headers = ['Warehouse Name', 'Location', 'Quantity (MT)', 'Loading Slips', 'Builties', 'Builty Numbers']
    for col, header in enumerate(wh_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    for wh in warehouse_dispatches:
        ws.cell(row=row, column=1, value=wh[0]).border = border
        ws.cell(row=row, column=2, value=wh[1]).border = border
        ws.cell(row=row, column=3, value=round(wh[2], 2)).border = border
        ws.cell(row=row, column=4, value=wh[3]).border = border
        ws.cell(row=row, column=5, value=wh[4]).border = border
        ws.cell(row=row, column=6, value=wh[5] or '-').border = border
        row += 1
    
    ws.cell(row=row, column=1, value="Total to Warehouses:").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(total_to_warehouses, 2)).font = Font(bold=True)
    row += 2
    
    # CGMF Section
    if cgmf_dispatches:
        ws.cell(row=row, column=1, value="DISPATCHED TO CGMF").font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        cgmf_headers = ['Society Name', 'District', 'Quantity (MT)', 'Loading Slips', 'Builties', 'Builty Numbers']
        for col, header in enumerate(cgmf_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        for cgmf in cgmf_dispatches:
            ws.cell(row=row, column=1, value=cgmf[0]).border = border
            ws.cell(row=row, column=2, value=cgmf[1]).border = border
            ws.cell(row=row, column=3, value=round(cgmf[3], 2)).border = border
            ws.cell(row=row, column=4, value=cgmf[4]).border = border
            ws.cell(row=row, column=5, value=cgmf[5]).border = border
            ws.cell(row=row, column=6, value=cgmf[6] or '-').border = border
            row += 1
        
        ws.cell(row=row, column=1, value="Total to CGMF:").font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(total_to_cgmf, 2)).font = Font(bold=True)
        row += 2
    
    # Summary Section
    ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = total_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    
    ws.cell(row=row, column=1, value="RR Quantity:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{round(rake_info[4], 2)} MT")
    row += 1
    ws.cell(row=row, column=1, value="Total Dispatched:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{round(total_dispatched, 2)} MT")
    row += 1
    ws.cell(row=row, column=1, value="Remaining:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{round(remaining, 2)} MT")
    
    # Adjust column widths for summary sheet
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 60
    # Wrap text in the Builty Numbers column (F) so long lists are readable
    for r in ws.iter_rows(min_col=6, max_col=6):
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # ---------------------------------------------------------------
    # Sheet 2: Builty Details (one row per builty, no truncation)
    # ---------------------------------------------------------------
    ws2 = wb.create_sheet(title="Builty Details")
    ws2['A1'] = f"BUILTY DETAILS - {rake_code}"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:K1')

    detail_headers = [
        'Builty No', 'Date', 'LR No', 'Recipient',
        'Sub Head', 'Truck No', 'Bags', 'MT',
        'Loading Point', 'Unloading Point', 'Freight (₹)'
    ]
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    detail_row = 4
    for b in all_builties:
        ws2.cell(row=detail_row, column=1,  value=b[0]).border = border
        ws2.cell(row=detail_row, column=2,  value=str(b[1]) if b[1] else '').border = border
        ws2.cell(row=detail_row, column=3,  value=b[2] or '').border = border
        ws2.cell(row=detail_row, column=4,  value=b[3]).border = border
        ws2.cell(row=detail_row, column=5,  value=b[4] or '').border = border
        ws2.cell(row=detail_row, column=6,  value=b[5] or '').border = border
        ws2.cell(row=detail_row, column=7,  value=b[6]).border = border
        ws2.cell(row=detail_row, column=8,  value=round(b[7], 2)).border = border
        ws2.cell(row=detail_row, column=9,  value=b[8] or '').border = border
        ws2.cell(row=detail_row, column=10, value=b[9] or '').border = border
        ws2.cell(row=detail_row, column=11, value=round(b[10], 2) if b[10] else 0).border = border
        detail_row += 1

    # Total row for detail sheet
    total_bags_sum   = sum(b[6] for b in all_builties) if all_builties else 0
    total_mt_sum     = round(sum(b[7] for b in all_builties), 2) if all_builties else 0
    total_freight_sum = round(sum((b[10] or 0) for b in all_builties), 2) if all_builties else 0
    rr_qty           = round(float(rake_info[4]), 2)
    difference       = round(rr_qty - total_mt_sum, 2)

    total_row_cell = ws2.cell(row=detail_row, column=1, value="TOTAL")
    total_row_cell.font = Font(bold=True, color="FFFFFF")
    total_row_cell.fill = total_fill
    total_row_cell.border = border
    for col in range(2, 12):
        ws2.cell(row=detail_row, column=col).fill = total_fill
        ws2.cell(row=detail_row, column=col).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=detail_row, column=col).border = border
    ws2.cell(row=detail_row, column=7, value=total_bags_sum)
    ws2.cell(row=detail_row, column=8, value=total_mt_sum)
    ws2.cell(row=detail_row, column=11, value=total_freight_sum)

    # RR Qty / Difference rows
    detail_row += 1
    ws2.cell(row=detail_row, column=7, value="RR Quantity (MT):").font = Font(bold=True)
    ws2.cell(row=detail_row, column=8, value=rr_qty).font = Font(bold=True)
    detail_row += 1
    diff_cell = ws2.cell(row=detail_row, column=7, value="Difference (RR − Total MT):")
    diff_cell.font = Font(bold=True)
    diff_val_cell = ws2.cell(row=detail_row, column=8, value=difference)
    diff_val_cell.font = Font(bold=True, color="FF0000" if difference > 0 else "22C55E")

    # Column widths for detail sheet
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 28
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 8
    ws2.column_dimensions['H'].width = 10
    ws2.column_dimensions['I'].width = 20
    ws2.column_dimensions['J'].width = 20
    ws2.column_dimensions['K'].width = 14
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'rake_{rake_code}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/download-total-summary-excel')
@login_required
def admin_download_total_summary_excel():
    """Download complete total summary as Excel file"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = Workbook()
    
    # === ACCOUNTS SHEET ===
    ws_accounts = wb.active
    ws_accounts.title = "Accounts Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    
    ws_accounts['A1'] = "ACCOUNTS (DEALERS/RETAILERS) SUMMARY"
    ws_accounts['A1'].font = Font(bold=True, size=14)
    ws_accounts.merge_cells('A1:F1')
    ws_accounts['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    acc_headers = ['#', 'Account Name', 'Type', 'Total Qty (MT)', 'Rakes', 'Loading Slips']
    for col, header in enumerate(acc_headers, 1):
        cell = ws_accounts.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    account_summary = db.execute_custom_query('''
        SELECT a.account_name, a.account_type, 
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN accounts a ON ls.account_id = a.id
        GROUP BY a.id, a.account_name, a.account_type
        ORDER BY total_quantity DESC
    ''')
    
    row = 5
    total_qty = 0
    for idx, item in enumerate(account_summary, 1):
        ws_accounts.cell(row=row, column=1, value=idx).border = border
        ws_accounts.cell(row=row, column=2, value=item[0]).border = border
        ws_accounts.cell(row=row, column=3, value=item[1]).border = border
        ws_accounts.cell(row=row, column=4, value=round(item[2], 2)).border = border
        ws_accounts.cell(row=row, column=5, value=item[3]).border = border
        ws_accounts.cell(row=row, column=6, value=item[4]).border = border
        total_qty += item[2]
        row += 1
    
    for col in range(1, 7):
        ws_accounts.cell(row=row, column=col).fill = total_fill
        ws_accounts.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws_accounts.cell(row=row, column=col).border = border
    ws_accounts.cell(row=row, column=1, value="TOTAL")
    ws_accounts.cell(row=row, column=4, value=round(total_qty, 2))
    
    ws_accounts.column_dimensions['A'].width = 5
    ws_accounts.column_dimensions['B'].width = 30
    ws_accounts.column_dimensions['C'].width = 12
    ws_accounts.column_dimensions['D'].width = 15
    ws_accounts.column_dimensions['E'].width = 10
    ws_accounts.column_dimensions['F'].width = 15
    
    # === CGMF SHEET ===
    ws_cgmf = wb.create_sheet("CGMF Summary")
    ws_cgmf['A1'] = "CGMF (CG MARKFED) SUMMARY"
    ws_cgmf['A1'].font = Font(bold=True, size=14)
    ws_cgmf.merge_cells('A1:G1')
    ws_cgmf['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    cgmf_headers = ['#', 'Society Name', 'District', 'Destination', 'Total Qty (MT)', 'Rakes', 'Loading Slips']
    for col, header in enumerate(cgmf_headers, 1):
        cell = ws_cgmf.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.border = border
    
    cgmf_summary = db.execute_custom_query('''
        SELECT c.society_name, c.district, c.destination,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN cgmf c ON ls.cgmf_id = c.id
        GROUP BY c.id, c.society_name, c.district
        ORDER BY total_quantity DESC
    ''')
    
    row = 5
    total_qty = 0
    for idx, item in enumerate(cgmf_summary, 1):
        ws_cgmf.cell(row=row, column=1, value=idx).border = border
        ws_cgmf.cell(row=row, column=2, value=item[0]).border = border
        ws_cgmf.cell(row=row, column=3, value=item[1]).border = border
        ws_cgmf.cell(row=row, column=4, value=item[2]).border = border
        ws_cgmf.cell(row=row, column=5, value=round(item[3], 2)).border = border
        ws_cgmf.cell(row=row, column=6, value=item[4]).border = border
        ws_cgmf.cell(row=row, column=7, value=item[5]).border = border
        total_qty += item[3]
        row += 1
    
    for col in range(1, 8):
        ws_cgmf.cell(row=row, column=col).fill = total_fill
        ws_cgmf.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws_cgmf.cell(row=row, column=col).border = border
    ws_cgmf.cell(row=row, column=1, value="TOTAL")
    ws_cgmf.cell(row=row, column=5, value=round(total_qty, 2))
    
    ws_cgmf.column_dimensions['A'].width = 5
    ws_cgmf.column_dimensions['B'].width = 30
    ws_cgmf.column_dimensions['C'].width = 15
    ws_cgmf.column_dimensions['D'].width = 15
    ws_cgmf.column_dimensions['E'].width = 15
    ws_cgmf.column_dimensions['F'].width = 10
    ws_cgmf.column_dimensions['G'].width = 15
    
    # === WAREHOUSE SHEET ===
    ws_wh = wb.create_sheet("Warehouse Summary")
    ws_wh['A1'] = "WAREHOUSE SUMMARY"
    ws_wh['A1'].font = Font(bold=True, size=14)
    ws_wh.merge_cells('A1:F1')
    ws_wh['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    wh_headers = ['#', 'Warehouse Name', 'Location', 'Total Qty (MT)', 'Rakes', 'Loading Slips']
    for col, header in enumerate(wh_headers, 1):
        cell = ws_wh.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        cell.border = border
    
    wh_summary = db.execute_custom_query('''
        SELECT w.warehouse_name, w.location,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN warehouses w ON ls.warehouse_id = w.id
        GROUP BY w.id, w.warehouse_name
        ORDER BY total_quantity DESC
    ''')
    
    row = 5
    total_qty = 0
    for idx, item in enumerate(wh_summary, 1):
        ws_wh.cell(row=row, column=1, value=idx).border = border
        ws_wh.cell(row=row, column=2, value=item[0]).border = border
        ws_wh.cell(row=row, column=3, value=item[1]).border = border
        ws_wh.cell(row=row, column=4, value=round(item[2], 2)).border = border
        ws_wh.cell(row=row, column=5, value=item[3]).border = border
        ws_wh.cell(row=row, column=6, value=item[4]).border = border
        total_qty += item[2]
        row += 1
    
    for col in range(1, 7):
        ws_wh.cell(row=row, column=col).fill = total_fill
        ws_wh.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws_wh.cell(row=row, column=col).border = border
    ws_wh.cell(row=row, column=1, value="TOTAL")
    ws_wh.cell(row=row, column=4, value=round(total_qty, 2))
    
    ws_wh.column_dimensions['A'].width = 5
    ws_wh.column_dimensions['B'].width = 25
    ws_wh.column_dimensions['C'].width = 20
    ws_wh.column_dimensions['D'].width = 15
    ws_wh.column_dimensions['E'].width = 10
    ws_wh.column_dimensions['F'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'total_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/download-accounts-summary-excel')
@login_required
def admin_download_accounts_summary_excel():
    """Download accounts summary as Excel file"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    
    ws['A1'] = "ACCOUNTS (DEALERS/RETAILERS) SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    headers = ['#', 'Account Name', 'Type', 'Total Qty (MT)', 'Rakes', 'Loading Slips']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    account_summary = db.execute_custom_query('''
        SELECT a.account_name, a.account_type, 
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN accounts a ON ls.account_id = a.id
        GROUP BY a.id, a.account_name, a.account_type
        ORDER BY total_quantity DESC
    ''')
    
    row = 5
    total_qty = 0
    for idx, item in enumerate(account_summary, 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item[0]).border = border
        ws.cell(row=row, column=3, value=item[1]).border = border
        ws.cell(row=row, column=4, value=round(item[2], 2)).border = border
        ws.cell(row=row, column=5, value=item[3]).border = border
        ws.cell(row=row, column=6, value=item[4]).border = border
        total_qty += item[2]
        row += 1
    
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=4, value=round(total_qty, 2))
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'accounts_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/download-cgmf-summary-excel')
@login_required
def admin_download_cgmf_summary_excel():
    """Download CGMF summary as Excel file"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CGMF Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    
    ws['A1'] = "CGMF (CG MARKFED) SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    headers = ['#', 'Society Name', 'District', 'Destination', 'Total Qty (MT)', 'Rakes', 'Loading Slips']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    cgmf_summary = db.execute_custom_query('''
        SELECT c.society_name, c.district, c.destination,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN cgmf c ON ls.cgmf_id = c.id
        GROUP BY c.id, c.society_name, c.district
        ORDER BY total_quantity DESC
    ''')
    
    row = 5
    total_qty = 0
    for idx, item in enumerate(cgmf_summary, 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item[0]).border = border
        ws.cell(row=row, column=3, value=item[1]).border = border
        ws.cell(row=row, column=4, value=item[2]).border = border
        ws.cell(row=row, column=5, value=round(item[3], 2)).border = border
        ws.cell(row=row, column=6, value=item[4]).border = border
        ws.cell(row=row, column=7, value=item[5]).border = border
        total_qty += item[3]
        row += 1
    
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=5, value=round(total_qty, 2))
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cgmf_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/download-warehouse-summary-excel')
@login_required
def admin_download_warehouse_summary_excel():
    """Download warehouse summary as Excel file"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    
    ws['A1'] = "WAREHOUSE SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    headers = ['#', 'Warehouse Name', 'Location', 'Total Qty (MT)', 'Rakes', 'Loading Slips']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    wh_summary = db.execute_custom_query('''
        SELECT w.warehouse_name, w.location,
               SUM(ls.quantity_mt) as total_quantity,
               COUNT(DISTINCT ls.rake_code) as rake_count,
               COUNT(ls.id) as slip_count
        FROM loading_slips ls
        JOIN warehouses w ON ls.warehouse_id = w.id
        GROUP BY w.id, w.warehouse_name
        ORDER BY total_quantity DESC
    ''')
    
    row = 5
    total_qty = 0
    for idx, item in enumerate(wh_summary, 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item[0]).border = border
        ws.cell(row=row, column=3, value=item[1]).border = border
        ws.cell(row=row, column=4, value=round(item[2], 2)).border = border
        ws.cell(row=row, column=5, value=item[3]).border = border
        ws.cell(row=row, column=6, value=item[4]).border = border
        total_qty += item[2]
        row += 1
    
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=4, value=round(total_qty, 2))
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'warehouse_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/manage-accounts', methods=['GET', 'POST'])
@login_required
def admin_manage_accounts():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        account_name = request.form.get('account_name')
        account_type = request.form.get('account_type')
        contact = request.form.get('contact')
        address = request.form.get('address')
        distance = request.form.get('distance', 0)
        try:
            distance = float(distance) if distance else 0
        except ValueError:
            distance = 0
        
        account_id = db.add_account(account_name, account_type, contact, address, distance)
        
        if account_id:
            flash(f'Account {account_name} added successfully!', 'success')
        else:
            flash('Error adding account', 'error')
    
    accounts = db.get_all_accounts()
    companies = db.get_all_companies()
    employees = db.get_all_employees()
    cgmf_list = db.get_all_cgmf()
    return render_template('admin/manage_accounts.html', accounts=accounts, companies=companies, employees=employees, cgmf_list=cgmf_list)

@app.route('/admin/add-company', methods=['POST'])
@login_required
def admin_add_company():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    company_name = request.form.get('company_name')
    company_code = request.form.get('company_code', '')
    contact_person = request.form.get('contact_person', '')
    mobile = request.form.get('mobile', '')
    address = request.form.get('address', '')
    distance = request.form.get('distance', 0)
    try:
        distance = float(distance) if distance else 0
    except ValueError:
        distance = 0
    
    company_id = db.add_company(company_name, company_code, contact_person, mobile, address, distance)
    
    if company_id:
        flash(f'Company {company_name} added successfully!', 'success')
    else:
        flash('Error adding company. Company may already exist.', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/add-employee', methods=['POST'])
@login_required
def admin_add_employee():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    employee_name = request.form.get('employee_name')
    employee_code = request.form.get('employee_code', '')
    mobile = request.form.get('mobile', '')
    designation = request.form.get('designation', '')
    
    employee_id = db.add_employee(employee_name, employee_code, mobile, designation)
    
    if employee_id:
        flash(f'Employee {employee_name} added successfully!', 'success')
    else:
        flash('Error adding employee.', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/add-cgmf', methods=['POST'])
@login_required
def admin_add_cgmf():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    district = request.form.get('district')
    destination = request.form.get('destination')
    society_name = request.form.get('society_name')
    contact = request.form.get('contact', '')
    distance = request.form.get('distance', 0)
    try:
        distance = float(distance) if distance else 0
    except ValueError:
        distance = 0
    
    cgmf_id = db.add_cgmf(district, destination, society_name, contact, distance)
    
    if cgmf_id:
        flash(f'CGMF Society "{society_name}" added successfully!', 'success')
    else:
        flash('Error adding CGMF society.', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/delete-account/<string:account_id>', methods=['POST'])
@login_required
def admin_delete_account(account_id):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    success, message = db.delete_account(account_id)
    
    if success:
        flash(message, 'success')
    else:
        flash(f'Error: {message}', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/edit-account/<string:account_id>', methods=['POST'])
@login_required
def admin_edit_account(account_id):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    account_name = request.form.get('account_name')
    account_type = request.form.get('account_type')
    contact = request.form.get('contact', '')
    address = request.form.get('address', '')
    distance = request.form.get('distance', 0)
    try:
        distance = float(distance) if distance else 0
    except ValueError:
        distance = 0
    
    success, message = db.update_account(account_id, account_name, account_type, contact, address, distance)
    
    if success:
        flash(f'Account "{account_name}" updated successfully!', 'success')
    else:
        flash(f'Error updating account: {message}', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/edit-company/<string:company_id>', methods=['POST'])
@login_required
def admin_edit_company(company_id):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    company_name = request.form.get('company_name')
    company_code = request.form.get('company_code', '')
    contact_person = request.form.get('contact_person', '')
    mobile = request.form.get('mobile', '')
    address = request.form.get('address', '')
    distance = request.form.get('distance', 0)
    try:
        distance = float(distance) if distance else 0
    except ValueError:
        distance = 0
    
    success, message = db.update_company(company_id, company_name, company_code, contact_person, mobile, address, distance)
    
    if success:
        flash(f'Company "{company_name}" updated successfully!', 'success')
    else:
        flash(f'Error updating company: {message}', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/edit-employee/<string:employee_id>', methods=['POST'])
@login_required
def admin_edit_employee(employee_id):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    employee_name = request.form.get('employee_name')
    employee_code = request.form.get('employee_code', '')
    mobile = request.form.get('mobile', '')
    designation = request.form.get('designation', '')
    
    success, message = db.update_employee(employee_id, employee_name, employee_code, mobile, designation)
    
    if success:
        flash(f'Employee "{employee_name}" updated successfully!', 'success')
    else:
        flash(f'Error updating employee: {message}', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/edit-cgmf/<string:cgmf_id>', methods=['POST'])
@login_required
def admin_edit_cgmf(cgmf_id):
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    district = request.form.get('district')
    destination = request.form.get('destination')
    society_name = request.form.get('society_name')
    contact = request.form.get('contact', '')
    distance = request.form.get('distance', 0)
    try:
        distance = float(distance) if distance else 0
    except ValueError:
        distance = 0
    
    success, message = db.update_cgmf(cgmf_id, district, destination, society_name, contact, distance)
    
    if success:
        flash(f'CGMF Society "{society_name}" updated successfully!', 'success')
    else:
        flash(f'Error updating CGMF: {message}', 'error')
    
    return redirect(url_for('admin_manage_accounts'))

@app.route('/admin/add-product', methods=['POST'])
@login_required
def admin_add_product():
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    product_name = request.form.get('product_name')
    product_code = request.form.get('product_code')
    product_type = request.form.get('product_type', 'Fertilizer')
    unit = request.form.get('unit', 'MT')
    unit_per_bag = float(request.form.get('unit_per_bag', 50.0))
    unit_type = request.form.get('unit_type', 'kg')
    description = request.form.get('description', '')
    
    product_id = db.add_product(product_name, product_code, product_type, unit, unit_per_bag, unit_type, description)
    
    if product_id:
        flash(f'Product {product_name} added successfully!', 'success')
    else:
        flash('Error adding product. Product may already exist.', 'error')
    
    return redirect(url_for('admin_add_rake'))


# ========== ADMIN Viewing Routes (All System Data) ==========

@app.route('/admin/all-loading-slips')
@login_required
def admin_all_loading_slips():
    """Admin view of all loading slips from all sources"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    loading_slips = db.get_all_loading_slips_with_status()
    accounts = db.get_all_accounts()
    warehouses = db.get_all_warehouses()
    cgmf_list = db.get_all_cgmf()
    rakes = db.get_all_rakes()
    trucks = db.get_all_trucks()
    return render_template('admin/all_loading_slips.html', loading_slips=loading_slips, accounts=accounts, warehouses=warehouses, cgmf_list=cgmf_list, rakes=rakes, trucks=trucks)

@app.route('/admin/all-builties')
@login_required
def admin_all_builties():
    """Admin view of all builties"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    builties = db.get_all_builties()
    accounts = db.get_all_accounts()
    warehouses = db.get_all_warehouses()
    cgmf_list = db.get_all_cgmf()
    rakes = db.get_all_rakes()
    trucks = db.get_all_trucks()
    return render_template('admin/all_builties.html', builties=builties, accounts=accounts, warehouses=warehouses, cgmf_list=cgmf_list, rakes=rakes, trucks=trucks)

@app.route('/admin/delete-builty/<string:builty_id>', methods=['POST'])
@login_required
def admin_delete_builty(builty_id):
    """Admin delete builty"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    delete_loading_slip = request.form.get('delete_loading_slip', 'false') == 'true'
    # Fetch builty info before deleting (for notification)
    builty_info = db.execute_custom_query('SELECT builty_number, rake_code FROM builty WHERE id = %s', (builty_id,))
    success, message = db.delete_builty(builty_id, delete_loading_slip)

    if success:
        flash(message, 'success')
        b_num = builty_info[0][0] if builty_info else builty_id
        b_rake = builty_info[0][1] if builty_info else 'N/A'
        send_formspree(
            subject=f'[FIMS] Builty {b_num} Deleted',
            message=(
                f'Admin {current_user.username} deleted builty {b_num} (Rake: {b_rake}).\n'
                f'Also deleted linked loading slip: {"Yes" if delete_loading_slip else "No"}'
            )
        )
    else:
        flash(message, 'error')

    return redirect(url_for('admin_all_builties'))

@app.route('/admin/delete-loading-slip/<string:slip_id>', methods=['POST'])
@login_required
def admin_delete_loading_slip(slip_id):
    """Admin delete loading slip"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    delete_builty = request.form.get('delete_builty', 'false') == 'true'
    # Fetch slip info before deleting (for notification)
    slip_info = db.execute_custom_query('SELECT slip_number, rake_code FROM loading_slips WHERE id = %s', (slip_id,))
    success, message = db.delete_loading_slip(slip_id, delete_builty)

    if success:
        flash(message, 'success')
        s_num = slip_info[0][0] if slip_info else slip_id
        s_rake = slip_info[0][1] if slip_info else 'N/A'
        send_formspree(
            subject=f'[FIMS] Loading Slip #{s_num} Deleted',
            message=(
                f'Admin {current_user.username} deleted loading slip #{s_num} (Rake: {s_rake}).\n'
                f'Also deleted linked builty: {"Yes" if delete_builty else "No"}'
            )
        )
    else:
        flash(message, 'error')

    return redirect(url_for('admin_all_loading_slips'))

@app.route('/admin/edit-loading-slip/<string:slip_id>', methods=['POST'])
@login_required
def admin_edit_loading_slip(slip_id):
    """Admin edit loading slip"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get destination from hidden field or manual entry
    destination = request.form.get('destination_final') or request.form.get('destination', '')
    account_id = request.form.get('account_id') or None
    warehouse_id = request.form.get('warehouse_id') or None
    cgmf_id = request.form.get('cgmf_id') or None
    
    quantity_bags = int(request.form.get('quantity_bags', 0))
    quantity_mt = float(request.form.get('quantity_mt', 0))
    goods_name = request.form.get('goods_name')
    date = request.form.get('date') or None  # New date field

    # Balance check: prevent increasing a slip beyond the rake's remaining capacity
    current = db.execute_custom_query(
        'SELECT quantity_mt, rake_code FROM loading_slips WHERE id = %s', (slip_id,)
    )
    if current:
        old_qty = float(current[0][0] or 0)
        slip_rake_code = current[0][1]
        delta = quantity_mt - old_qty
        if delta > 0 and slip_rake_code and slip_rake_code != 'WAREHOUSE':
            balance = db.get_rake_balance(slip_rake_code)
            if balance and delta > balance['remaining']:
                flash(
                    f'Cannot increase quantity: would exceed rake balance. '
                    f'Available: {balance["remaining"]:.2f} MT, '
                    f'Increase requested: {delta:.2f} MT', 'error'
                )
                return redirect(url_for('admin_all_loading_slips'))

    success, message = db.update_loading_slip(slip_id, destination, quantity_bags, quantity_mt, goods_name, account_id, warehouse_id, cgmf_id, date)
    
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    
    return redirect(url_for('admin_all_loading_slips'))

@app.route('/admin/edit-builty/<string:builty_id>', methods=['POST'])
@login_required
def admin_edit_builty(builty_id):
    """Admin edit builty"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get unloading point from hidden field or manual entry
    unloading_point = request.form.get('unloading_point_final') or request.form.get('unloading_point', '')
    account_id = request.form.get('account_id') or None
    warehouse_id = request.form.get('warehouse_id') or None
    cgmf_id = request.form.get('cgmf_id') or None
    
    number_of_bags = int(request.form.get('number_of_bags', 0))
    quantity_mt = float(request.form.get('quantity_mt', 0))
    rate_per_mt = float(request.form.get('rate_per_mt', 0))
    total_freight = float(request.form.get('total_freight', 0))
    advance = float(request.form.get('advance', 0))
    to_pay = float(request.form.get('to_pay', 0))
    date = request.form.get('date') or None  # New date field
    
    success, message = db.update_builty(builty_id, unloading_point, number_of_bags, quantity_mt, rate_per_mt, total_freight, advance, to_pay, account_id, warehouse_id, cgmf_id, date)
    
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    
    return redirect(url_for('admin_all_builties'))

@app.route('/admin/all-ebills')
@login_required
def admin_all_ebills():
    """Admin view of all e-bills"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    ebills = db.get_all_ebills()
    return render_template('admin/all_ebills.html', ebills=ebills)

@app.route('/admin/warehouse-transactions')
@login_required
def admin_warehouse_transactions():
    """Admin view of all warehouse stock IN and OUT transactions - side by side format"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get filter parameters
    selected_product = request.args.get('product', 'all')
    selected_warehouse = request.args.get('warehouse', 'all')
    
    # Get filter options
    products = db.get_all_products()
    warehouses = db.get_all_warehouses()
    
    # Build WHERE clause for filtering
    conditions = ["1=1"]
    params = []
    
    if selected_product != 'all':
        conditions.append("p.id = ?")
        params.append(selected_product)
    
    if selected_warehouse != 'all':
        conditions.append("w.id = ?")
        params.append(selected_warehouse)
    
    where_clause = " AND ".join(conditions)
    
    # Get Stock IN transactions (excluding allotments)
    stock_in = db.execute_custom_query(f'''
        SELECT ws.date, p.product_name, ws.quantity_mt, c.company_name, 
               COALESCE(a.account_name, 'N/A') as employee,
               w.warehouse_name
        FROM warehouse_stock ws
        LEFT JOIN warehouses w ON ws.warehouse_id = w.id
        LEFT JOIN products p ON ws.product_id = p.id
        LEFT JOIN companies c ON ws.company_id = c.id
        LEFT JOIN accounts a ON ws.account_id = a.id
        WHERE ws.transaction_type = 'IN' AND (ws.source_type IS NULL OR ws.source_type::text != 'allotment') AND {where_clause}
        ORDER BY ws.date DESC, ws.created_at DESC
    ''', tuple(params)) or []
    
    # Get Stock OUT transactions (excluding allotments)
    stock_out = db.execute_custom_query(f'''
        SELECT ws.date, p.product_name, ws.quantity_mt, c.company_name,
               COALESCE(a.account_name, 'N/A') as employee,
               w.warehouse_name
        FROM warehouse_stock ws
        LEFT JOIN warehouses w ON ws.warehouse_id = w.id
        LEFT JOIN products p ON ws.product_id = p.id
        LEFT JOIN companies c ON ws.company_id = c.id
        LEFT JOIN accounts a ON ws.account_id = a.id
        WHERE ws.transaction_type = 'OUT' AND (ws.source_type IS NULL OR ws.source_type::text != 'allotment') AND {where_clause}
        ORDER BY ws.date DESC, ws.created_at DESC
    ''', tuple(params)) or []
    
    # Calculate totals
    total_in = sum(t[2] for t in stock_in) if stock_in else 0
    total_out = sum(t[2] for t in stock_out) if stock_out else 0
    
    return render_template('admin/warehouse_transactions.html',
                         stock_in=stock_in,
                         stock_out=stock_out,
                         total_in=total_in,
                         total_out=total_out,
                         products=products,
                         warehouses=warehouses,
                         selected_product=selected_product,
                         selected_warehouse=selected_warehouse)

@app.route('/admin/warehouse-summary')
@login_required
def admin_warehouse_summary():
    """Admin view of warehouse stock summary - redesigned with detailed transaction view"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get filter parameters
    selected_warehouse = request.args.get('warehouse', 'all')
    selected_company = request.args.get('company', 'all')
    selected_product = request.args.get('product', 'all')
    selected_account = request.args.get('account', 'all')
    date_filter = request.args.get('date_filter', 'all')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Get all data for filters
    warehouses = db.get_all_warehouses()
    companies = db.get_all_companies()
    products = db.get_all_products()
    accounts = db.get_all_accounts()
    
    # Build query conditions
    conditions = ["1=1"]
    params = []
    
    if selected_warehouse != 'all':
        conditions.append("w.id = ?")
        params.append(selected_warehouse)
    
    if selected_company != 'all':
        conditions.append("c.id = ?")
        params.append(selected_company)
    
    if selected_product != 'all':
        conditions.append("p.id = ?")
        params.append(selected_product)
    
    if selected_account != 'all':
        conditions.append("a.id = ?")
        params.append(selected_account)
    
    # Date filtering
    from datetime import datetime, timedelta
    if date_filter == 'today':
        conditions.append("DATE(ws.date) = ?")
        params.append(datetime.now().strftime('%Y-%m-%d'))
    elif date_filter == 'week':
        week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        conditions.append("DATE(ws.date) >= ?")
        params.append(week_ago)
    elif date_filter == 'month':
        month_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        conditions.append("DATE(ws.date) >= ?")
        params.append(month_ago)
    elif date_filter == 'custom' and start_date and end_date:
        conditions.append("DATE(ws.date) >= ? AND DATE(ws.date) <= ?")
        params.extend([start_date, end_date])
    
    where_clause = " AND ".join(conditions)
    
    # Get grouped data: Product -> Company -> Warehouse breakdown (excluding allotments)
    product_summary = db.execute_custom_query(f'''
        SELECT p.product_name, 
               SUM(ws.quantity_mt) as total_qty,
               c.company_name,
               SUM(ws.quantity_mt) as company_qty,
               w.warehouse_name
        FROM warehouse_stock ws
        LEFT JOIN warehouses w ON ws.warehouse_id = w.id
        LEFT JOIN companies c ON ws.company_id = c.id
        LEFT JOIN products p ON ws.product_id = p.id
        LEFT JOIN accounts a ON ws.account_id = a.id
        WHERE {where_clause} AND ws.transaction_type = 'IN' AND (ws.source_type IS NULL OR ws.source_type::text != 'allotment')
        GROUP BY p.product_name, c.company_name, w.warehouse_name
        ORDER BY p.product_name, c.company_name, w.warehouse_name
    ''', tuple(params)) or []
    
    # Organize data into grouped structure
    from collections import OrderedDict
    grouped_data = OrderedDict()
    for row in product_summary:
        product = row[0] or 'Unknown'
        company = row[2] or 'Unknown'
        qty = row[3] or 0
        warehouse = row[4] or 'Unknown'
        
        if product not in grouped_data:
            grouped_data[product] = {'total': 0, 'companies': []}
        
        grouped_data[product]['total'] += qty
        grouped_data[product]['companies'].append({
            'company': company,
            'qty': qty,
            'warehouse': warehouse
        })
    
    # Calculate grand total
    total_quantity = sum(data['total'] for data in grouped_data.values())
    
    return render_template('admin/warehouse_summary.html',
                         grouped_data=grouped_data,
                         total_quantity=total_quantity,
                         warehouses=warehouses,
                         companies=companies,
                         products=products,
                         accounts=accounts,
                         selected_warehouse=selected_warehouse,
                         selected_company=selected_company,
                         selected_product=selected_product,
                         selected_account=selected_account,
                         date_filter=date_filter,
                         start_date=start_date,
                         end_date=end_date)

# ========== Logistic Bill Routes ==========

@app.route('/admin/logistic-bill')
@login_required
def admin_logistic_bill():
    """Logistic Bill page with Rake Bill and Warehouse Bill sections - OPTIMIZED"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    try:
        rakes = db.get_all_rakes() or []
        warehouses = db.get_all_warehouses() or []
        companies = db.get_all_companies() or []
        
        # Get filter parameters
        selected_company = request.args.get('company', 'all')
        selected_rake = request.args.get('rake', 'all')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Use optimized single-query method for bill summary
        bill_summary = db.get_logistic_bill_summary_optimized(selected_company, selected_rake, date_from, date_to)
        
        # Storage and transport data will be loaded via AJAX to reduce initial load time
        storage_data = []
        transport_data = []
        
    except Exception as e:
        print(f"Error loading logistic bill data: {e}")
        rakes = []
        warehouses = []
        storage_data = []
        transport_data = []
        bill_summary = []
        companies = []
        selected_company = 'all'
        selected_rake = 'all'
        date_from = ''
        date_to = ''
        flash('Error loading some data. Please try again.', 'warning')
    
    return render_template('admin/logistic_bill.html',
                         rakes=rakes,
                         warehouses=warehouses,
                         storage_data=storage_data,
                         transport_data=transport_data,
                         bill_summary=bill_summary,
                         companies=companies,
                         selected_company=selected_company,
                         selected_rake=selected_rake,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/admin/logistic-bill/rake-data/<path:rake_code>')
@login_required
def admin_logistic_bill_rake_data(rake_code):
    """API to get rake transport data for logistic bill"""
    if current_user.role != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    try:
        data = db.get_rake_transport_data(rake_code)
    except Exception as e:
        print(f"Error fetching rake data: {e}")
        data = []
    return jsonify(data)

@app.route('/admin/logistic-bill/warehouse-data')
@app.route('/admin/logistic-bill/warehouse-data/<string:warehouse_id>')
@login_required
def admin_logistic_bill_warehouse_data(warehouse_id=None):
    """API to get warehouse storage and transport data for logistic bill"""
    if current_user.role != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        storage_data = db.get_warehouse_storage_data(warehouse_id) or []
        transport_data = db.get_warehouse_transport_data(warehouse_id) or []
    except Exception as e:
        print(f"Error fetching warehouse data: {e}")
        storage_data = []
        transport_data = []
    
    # Convert to JSON-friendly format
    storage_list = []
    for item in storage_data:
        storage_list.append({
            'date': str(item[0]) if item[0] else '',
            'company': item[1] or '',
            'product': item[2] or '',
            'quantity': item[3] or 0,
            'stock_id': item[4]
        })
    
    transport_list = []
    for item in transport_data:
        transport_list.append({
            'date': str(item[0]) if item[0] else '',
            'truck': item[1] or '',
            'account': item[2] or '',
            'product': item[3] or '',
            'quantity': item[4] or 0,
            'distance': item[5] or 0,
            'stock_id': item[6]
        })
    
    return jsonify({
        'storage': storage_list,
        'transport': transport_list
    })

@app.route('/admin/logistic-bill/download-rake-excel', methods=['POST'])
@login_required
def admin_download_rake_excel():
    """Download Rake Bill as Excel file"""
    if current_user.role != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    data = request.get_json()
    rake_code = data.get('rake_code')
    transport_data = data.get('transport_data', [])
    handling_data = data.get('handling_data', [])
    totals = data.get('totals', {})
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rake Bill"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    
    # Title
    ws['A1'] = f"RAKE BILL - {rake_code}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Rake Transport Section
    ws['A3'] = "1. RAKE TRANSPORT"
    ws['A3'].font = Font(bold=True, size=12)
    
    # Transport Headers
    transport_headers = ['Type', 'Account/Destination', 'QT (MT)', 'KM Distance', 'Rate (₹)', 'Total (₹)']
    for col, header in enumerate(transport_headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Transport Data
    row = 5
    for item in transport_data:
        ws.cell(row=row, column=1, value=item.get('dest_type', '')).border = border
        ws.cell(row=row, column=2, value=item.get('dest_name', '')).border = border
        ws.cell(row=row, column=3, value=item.get('quantity', 0)).border = border
        ws.cell(row=row, column=4, value=item.get('distance', 0)).border = border
        ws.cell(row=row, column=5, value=item.get('rate', 0)).border = border
        ws.cell(row=row, column=6, value=item.get('total', 0)).border = border
        row += 1
    
    # Transport Total
    ws.cell(row=row, column=5, value="Transport Total:").font = Font(bold=True)
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=6, value=totals.get('transport_total', 0)).border = border
    ws.cell(row=row, column=6).font = Font(bold=True)
    row += 2
    
    # Rake Handling Section
    ws.cell(row=row, column=1, value="2. RAKE HANDLING").font = Font(bold=True, size=12)
    row += 1
    
    # Handling Headers
    handling_headers = ['Handling Situation', 'QT (MT)', 'Rate (₹)', 'Total (₹)']
    for col, header in enumerate(handling_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Handling Data
    for item in handling_data:
        ws.cell(row=row, column=1, value=item.get('situation', '')).border = border
        ws.cell(row=row, column=2, value=item.get('quantity', 0)).border = border
        ws.cell(row=row, column=3, value=item.get('rate', 0)).border = border
        ws.cell(row=row, column=4, value=item.get('total', 0)).border = border
        row += 1
    
    # Handling Total
    ws.cell(row=row, column=3, value="Handling Total:").font = Font(bold=True)
    ws.cell(row=row, column=3).border = border
    ws.cell(row=row, column=4, value=totals.get('handling_total', 0)).border = border
    ws.cell(row=row, column=4).font = Font(bold=True)
    row += 2
    
    # Grand Total
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=14)
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=6, value=totals.get('grand_total', 0))
    ws.cell(row=row, column=6).fill = total_fill
    ws.cell(row=row, column=6).font = Font(bold=True, color="FFFFFF", size=14)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Rake_Bill_{rake_code}.xlsx'
    )

@app.route('/admin/logistic-bill/download-warehouse-excel', methods=['POST'])
@login_required
def admin_download_warehouse_excel():
    """Download Warehouse Bill as Excel file"""
    if current_user.role != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    data = request.get_json()
    warehouse_name = data.get('warehouse_name', 'All Warehouses')
    storage_data = data.get('storage_data', [])
    transport_data = data.get('transport_data', [])
    totals = data.get('totals', {})
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Bill"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    
    # Title
    ws['A1'] = f"WAREHOUSE BILL - {warehouse_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Storage Section
    ws['A3'] = "1. STORAGE (Stock In)"
    ws['A3'].font = Font(bold=True, size=12)
    
    # Storage Headers
    storage_headers = ['Date', 'Company', 'Product', 'QT (MT)', 'Rate (₹)', 'Total (₹)']
    for col, header in enumerate(storage_headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Storage Data
    row = 5
    for item in storage_data:
        ws.cell(row=row, column=1, value=item.get('date', '')).border = border
        ws.cell(row=row, column=2, value=item.get('company', '')).border = border
        ws.cell(row=row, column=3, value=item.get('product', '')).border = border
        ws.cell(row=row, column=4, value=item.get('quantity', 0)).border = border
        ws.cell(row=row, column=5, value=item.get('rate', 0)).border = border
        ws.cell(row=row, column=6, value=item.get('total', 0)).border = border
        row += 1
    
    # Storage Total
    ws.cell(row=row, column=5, value="Storage Total:").font = Font(bold=True)
    ws.cell(row=row, column=5).border = border
    ws.cell(row=row, column=6, value=totals.get('storage_total', 0)).border = border
    ws.cell(row=row, column=6).font = Font(bold=True)
    row += 2
    
    # Transportation Section
    ws.cell(row=row, column=1, value="2. TRANSPORTATION (Stock Out - Builty Wise)").font = Font(bold=True, size=12)
    row += 1
    
    # Transport Headers
    transport_headers = ['Date', 'Truck', 'Account', 'Product', 'QT (MT)', 'KM', 'Rate (₹)', 'Total (₹)']
    for col, header in enumerate(transport_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Transport Data
    for item in transport_data:
        ws.cell(row=row, column=1, value=item.get('date', '')).border = border
        ws.cell(row=row, column=2, value=item.get('truck', '')).border = border
        ws.cell(row=row, column=3, value=item.get('account', '')).border = border
        ws.cell(row=row, column=4, value=item.get('product', '')).border = border
        ws.cell(row=row, column=5, value=item.get('quantity', 0)).border = border
        ws.cell(row=row, column=6, value=item.get('distance', 0)).border = border
        ws.cell(row=row, column=7, value=item.get('rate', 0)).border = border
        ws.cell(row=row, column=8, value=item.get('total', 0)).border = border
        row += 1
    
    # Transport Total
    ws.cell(row=row, column=7, value="Transport Total:").font = Font(bold=True)
    ws.cell(row=row, column=7).border = border
    ws.cell(row=row, column=8, value=totals.get('transport_total', 0)).border = border
    ws.cell(row=row, column=8).font = Font(bold=True)
    row += 2
    
    # Grand Total
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=14)
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=8, value=totals.get('grand_total', 0))
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).font = Font(bold=True, color="FFFFFF", size=14)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Warehouse_Bill_{warehouse_name.replace(" ", "_")}.xlsx'
    )

@app.route('/admin/download-bill-summary-excel')
@login_required
def admin_download_bill_summary_excel():
    """Download Bill Summary as Excel file"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    # Get filter parameters
    selected_company = request.args.get('company', 'all')
    selected_rake = request.args.get('rake', 'all')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill Summary"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    # Title
    ws['A1'] = "BILL SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Subtitle with filters
    filter_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from or date_to:
        filter_text += f" | Date Range: {date_from or 'Start'} to {date_to or 'End'}"
    if selected_company != 'all':
        filter_text += f" | Company: {selected_company}"
    if selected_rake != 'all':
        filter_text += f" | Rake: {selected_rake}"
    ws['A2'] = filter_text
    
    # Headers
    headers = ['#', 'Rake', 'Company', 'Total Stock (MT)', 'Date', 'Bill Amount (₹)', 'Received (₹)', 'Left (₹)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Get filtered bill summary data
    bill_summary = db.get_logistic_bill_summary_optimized(selected_company, selected_rake, date_from, date_to)
    
    row = 5
    total_stock = 0
    total_bill = 0
    total_received = 0
    
    for idx, bill in enumerate(bill_summary, 1):
        rake_code = bill['rake_code']
        company_name = bill['company_name']
        rr_quantity = bill['total_stock'] or 0
        date = bill['date'] or ''
        bill_amount = bill['bill_amount'] or 0
        received_payment = bill['received_payment'] or 0
        left = bill_amount - received_payment
        
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=rake_code).border = border
        ws.cell(row=row, column=3, value=company_name).border = border
        ws.cell(row=row, column=4, value=round(rr_quantity, 2)).border = border
        ws.cell(row=row, column=5, value=str(date)).border = border
        ws.cell(row=row, column=6, value=round(bill_amount, 2)).border = border
        ws.cell(row=row, column=7, value=round(received_payment, 2)).border = border
        ws.cell(row=row, column=8, value=round(left, 2)).border = border
        
        total_stock += rr_quantity
        total_bill += bill_amount
        total_received += received_payment
        row += 1
    
    # Total row
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=col).border = border
    
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=4, value=round(total_stock, 2))
    ws.cell(row=row, column=6, value=round(total_bill, 2))
    ws.cell(row=row, column=7, value=round(total_received, 2))
    ws.cell(row=row, column=8, value=round(total_bill - total_received, 2))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Bill_Summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/admin/save-rake-bill-payment', methods=['POST'])
@login_required
def admin_save_rake_bill_payment():
    """Save rake bill payment information"""
    if current_user.role != 'Admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        rake_code = data.get('rake_code')
        total_bill_amount = float(data.get('total_bill_amount', 0))
        received_amount = float(data.get('received_amount', 0))
        
        if not rake_code:
            return jsonify({'success': False, 'error': 'Rake code is required'}), 400
        
        success = db.save_rake_bill_payment(rake_code, total_bill_amount, received_amount, current_user.username)
        
        if success:
            return jsonify({'success': True, 'message': 'Payment saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save payment'}), 500
    except Exception as e:
        print(f"Error saving rake bill payment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/get-rake-bill-payment/<path:rake_code>')
@login_required
def admin_get_rake_bill_payment(rake_code):
    """Get rake bill payment information"""
    if current_user.role != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    try:
        payment_data = db.get_rake_bill_payment(rake_code)
        if payment_data:
            return jsonify(payment_data)
        else:
            return jsonify({'error': 'No payment data found'}), 404
    except Exception as e:
        print(f"Error getting rake bill payment: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/warehouse-account-stock/<string:warehouse_id>')
@login_required
def api_warehouse_account_stock(warehouse_id):
    """API to get account-wise stock breakdown for a warehouse"""
    try:
        # Detect PK column names dynamically (same as warehouse_dashboard)
        def _pk_col(table_name, legacy_pk):
            try:
                cols = db.execute_custom_query(
                    f"SELECT name FROM pragma_table_info('{table_name}')"
                ) or []
                names = [c[0] for c in cols if c and len(c) > 0]
                if names:
                    return 'id' if 'id' in names else legacy_pk
            except Exception:
                pass
            try:
                cols = db.execute_custom_query(
                    f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'"
                ) or []
                names = [c[0] for c in cols if c and len(c) > 0]
                if names:
                    return 'id' if 'id' in names else legacy_pk
            except Exception:
                pass
            return legacy_pk

        account_pk = _pk_col('accounts', 'account_id')
        cgmf_pk = _pk_col('cgmf', 'cgmf_id')
        company_pk = _pk_col('companies', 'company_id')
        product_pk = _pk_col('products', 'product_id')

        account_stock = db.execute_custom_query(f'''
            SELECT a.{account_pk}, a.account_name,
                   CASE
                       WHEN LOWER(COALESCE(a.account_type, '')) IN ('company', 'payal') THEN 'Payal'
                       WHEN LOWER(COALESCE(a.account_type, '')) = 'dealer' THEN 'Dealer'
                       WHEN LOWER(COALESCE(a.account_type, '')) = 'retailer' THEN 'Retailer'
                       ELSE COALESCE(a.account_type, 'Account')
                   END AS account_type,
                   COALESCE(p.product_name, 'N/A') AS product_name,
                   COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) AS total_stock
            FROM warehouse_stock ws
            JOIN accounts a ON ws.account_id = a.{account_pk}
            LEFT JOIN products p ON ws.product_id = p.{product_pk}
            WHERE ws.warehouse_id = ? AND ws.account_id IS NOT NULL
            GROUP BY a.{account_pk}, a.account_name,
                     CASE
                         WHEN LOWER(COALESCE(a.account_type, '')) IN ('company', 'payal') THEN 'Payal'
                         WHEN LOWER(COALESCE(a.account_type, '')) = 'dealer' THEN 'Dealer'
                         WHEN LOWER(COALESCE(a.account_type, '')) = 'retailer' THEN 'Retailer'
                         ELSE COALESCE(a.account_type, 'Account')
                     END,
                     COALESCE(p.product_name, 'N/A')
            HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ''', (warehouse_id,)) or []

        cgmf_stock = db.execute_custom_query(f'''
            SELECT c.{cgmf_pk}, c.society_name, 'CGMF' AS account_type,
                   COALESCE(p.product_name, 'N/A') AS product_name,
                   COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) AS total_stock
            FROM warehouse_stock ws
            JOIN cgmf c ON ws.cgmf_id = c.{cgmf_pk}
            LEFT JOIN products p ON ws.product_id = p.{product_pk}
            WHERE ws.warehouse_id = ? AND ws.cgmf_id IS NOT NULL
            GROUP BY c.{cgmf_pk}, c.society_name, COALESCE(p.product_name, 'N/A')
            HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ''', (warehouse_id,)) or []

        company_stock = db.execute_custom_query(f'''
            SELECT c.{company_pk}, c.company_name, 'Company' AS account_type,
                   COALESCE(p.product_name, 'N/A') AS product_name,
                   COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) AS total_stock
            FROM warehouse_stock ws
            JOIN companies c ON ws.company_id = c.{company_pk}
            LEFT JOIN products p ON ws.product_id = p.{product_pk}
            WHERE ws.warehouse_id = ?
              AND ws.company_id IS NOT NULL
              AND ws.account_id IS NULL
              AND ws.cgmf_id IS NULL
            GROUP BY c.{company_pk}, c.company_name, COALESCE(p.product_name, 'N/A')
            HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ''', (warehouse_id,)) or []

        result = []
        for row in account_stock + cgmf_stock + company_stock:
            result.append({
                'account_id': row[0],
                'account_name': row[1],
                'account_type': row[2],
                'product_name': row[3],
                'total_stock': float(row[4] or 0)
            })

        return jsonify(result)
    except Exception as e:
        print(f"Error in api_warehouse_account_stock: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/edit-warehouse-stock', methods=['GET', 'POST'])
@login_required
def admin_edit_warehouse_stock():
    """Admin can view and adjust stock by Account/Company/CGMF"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        warehouse_id = request.form.get('warehouse_id')
        entity_type = request.form.get('entity_type')  # account, cgmf, company
        entity_id = request.form.get('entity_id')
        adjustment = float(request.form.get('adjustment', 0))
        reason = request.form.get('reason', '')
        
        if adjustment != 0:
            # Create an adjustment entry
            transaction_type = 'IN' if adjustment > 0 else 'OUT'
            qty = abs(adjustment)
            
            if entity_type == 'account':
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, account_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, ?, CURRENT_DATE, ?, 'adjustment')
                ''', (warehouse_id, entity_id, qty, transaction_type, f'Admin adjustment: {reason}'))
            elif entity_type == 'cgmf':
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, cgmf_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, ?, CURRENT_DATE, ?, 'adjustment')
                ''', (warehouse_id, entity_id, qty, transaction_type, f'Admin adjustment: {reason}'))
            elif entity_type == 'company':
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, company_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, ?, CURRENT_DATE, ?, 'adjustment')
                ''', (warehouse_id, entity_id, qty, transaction_type, f'Admin adjustment: {reason}'))
            
            flash(f'Stock adjusted by {adjustment:+.2f} MT successfully!', 'success')
        else:
            flash('No adjustment made (quantity was 0)', 'warning')
        
        return redirect(url_for('admin_edit_warehouse_stock'))
    
    # Get stock by Account
    account_stock = db.execute_custom_query('''
        SELECT w.id, w.warehouse_name, a.id, a.account_name, a.account_type,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.id
        JOIN accounts a ON ws.account_id = a.id
        WHERE ws.account_id IS NOT NULL
        GROUP BY w.id, a.id
        HAVING balance != 0
        ORDER BY w.warehouse_name, a.account_name
    ''') or []
    
    # Get stock by CGMF
    cgmf_stock = db.execute_custom_query('''
        SELECT w.id, w.warehouse_name, c.id, c.society_name, 'CGMF' as type,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.id
        JOIN cgmf c ON ws.cgmf_id = c.id
        WHERE ws.cgmf_id IS NOT NULL
        GROUP BY w.id, c.id
        HAVING balance != 0
        ORDER BY w.warehouse_name, c.society_name
    ''') or []
    
    # Get stock by Company (direct company stock, not via accounts)
    company_stock = db.execute_custom_query('''
        SELECT w.id, w.warehouse_name, c.id, c.company_name, 'Company' as type,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.id
        JOIN companies c ON ws.company_id = c.id
        WHERE ws.company_id IS NOT NULL AND ws.account_id IS NULL AND ws.cgmf_id IS NULL
        GROUP BY w.id, c.id
        HAVING balance != 0
        ORDER BY w.warehouse_name, c.company_name
    ''') or []
    
    # Combine all with entity_type marker
    all_stock = []
    for s in account_stock:
        all_stock.append({
            'warehouse_id': s[0], 'warehouse_name': s[1], 
            'entity_id': s[2], 'entity_name': s[3], 'entity_type': 'account',
            'type_display': s[4], 'balance': s[5]
        })
    for s in cgmf_stock:
        all_stock.append({
            'warehouse_id': s[0], 'warehouse_name': s[1],
            'entity_id': s[2], 'entity_name': s[3], 'entity_type': 'cgmf',
            'type_display': 'CGMF', 'balance': s[5]
        })
    for s in company_stock:
        all_stock.append({
            'warehouse_id': s[0], 'warehouse_name': s[1],
            'entity_id': s[2], 'entity_name': s[3], 'entity_type': 'company',
            'type_display': 'Company', 'balance': s[5]
        })
    
    return render_template('admin/edit_warehouse_stock.html', all_stock=all_stock)

@app.route('/admin/download-eway-bill/<filename>')
@login_required
def admin_download_eway_bill(filename):
    """Admin can also download eway bill PDFs"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'eway_bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('admin_all_ebills'))

@app.route('/admin/download-bill/<filename>')
@login_required
def admin_download_bill(filename):
    """Admin can also download bill PDFs"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('admin_all_ebills'))

@app.route('/admin/manage-warehouses', methods=['GET'])
@login_required
def admin_manage_warehouses():
    """Admin view to manage warehouses"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    warehouses = db.get_all_warehouses()
    return render_template('admin/manage_warehouses.html', warehouses=warehouses)

@app.route('/admin/add-warehouse', methods=['POST'])
@login_required
def admin_add_warehouse():
    """Admin can add new warehouse"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    warehouse_name = request.form.get('warehouse_name')
    location = request.form.get('location', '')
    capacity = float(request.form.get('capacity', 0))
    distance = request.form.get('distance', 0)
    try:
        distance = float(distance) if distance else 0
    except ValueError:
        distance = 0
    
    warehouse_id = db.add_warehouse(warehouse_name, location, capacity, distance)
    
    if warehouse_id:
        flash(f'Warehouse "{warehouse_name}" added successfully!', 'success')
    else:
        flash('Error adding warehouse.', 'error')
    
    return redirect(url_for('admin_manage_warehouses'))

@app.route('/admin/edit-warehouse/<string:warehouse_id>', methods=['POST'])
@login_required
def admin_edit_warehouse(warehouse_id):
    """Admin can edit warehouse"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    warehouse_name = request.form.get('warehouse_name')
    location = request.form.get('location', '')
    capacity = float(request.form.get('capacity', 0))
    
    success = db.update_warehouse(warehouse_id, warehouse_name, location, capacity)
    
    if success:
        flash(f'Warehouse "{warehouse_name}" updated successfully!', 'success')
    else:
        flash('Error updating warehouse.', 'error')
    
    return redirect(url_for('admin_manage_warehouses'))

@app.route('/admin/delete-warehouse/<string:warehouse_id>', methods=['POST'])
@login_required
def admin_delete_warehouse(warehouse_id):
    """Admin can delete warehouse"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    success, message = db.delete_warehouse(warehouse_id)
    
    if success:
        flash(message, 'success')
    else:
        flash(f'Error: {message}', 'error')
    
    return redirect(url_for('admin_manage_warehouses'))

@app.route('/admin/print-loading-slip/<string:slip_id>')
@login_required
def admin_print_loading_slip(slip_id):
    """Admin can print any loading slip"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get loading slip details
    slip = db.execute_custom_query('''
        SELECT ls.id, ls.rake_code, ls.slip_number, ls.loading_point_name,
               ls.destination, ls.quantity_bags, ls.quantity_mt, ls.goods_name,
               ls.truck_driver, ls.truck_owner, ls.mobile_number_1, ls.mobile_number_2,
               ls.wagon_number, ls.created_at, t.truck_number,
               COALESCE(a.account_name, w.warehouse_name, cg.society_name) as destination_name,
               COALESCE(a.account_type::text, CASE WHEN cg.id IS NOT NULL THEN 'CGMF' ELSE 'Warehouse' END) as destination_type
        FROM loading_slips ls
        LEFT JOIN trucks t ON ls.truck_id = t.id
        LEFT JOIN accounts a ON ls.account_id = a.id
        LEFT JOIN warehouses w ON ls.warehouse_id = w.id
        LEFT JOIN cgmf cg ON ls.cgmf_id = cg.id
        WHERE ls.id = ?
    ''', (slip_id,))
    
    if not slip:
        flash('Loading slip not found', 'error')
        return redirect(url_for('admin_all_loading_slips'))
    
    return render_template('print_loading_slip.html', slip=slip[0])

@app.route('/admin/print-builty/<string:builty_id>')
@login_required
def admin_print_builty(builty_id):
    """Admin can print any builty"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get builty details with all related information
    builty = db.get_builty_by_id(builty_id)
    
    if not builty:
        flash('Builty not found', 'error')
        return redirect(url_for('admin_all_builties'))
    
    # Convert to dict for easier template access
    # Explicit column order from get_builty_by_id query
    builty_dict = {
        'builty_id': builty[0],
        'builty_number': builty[1],
        'rake_code': builty[2],
        'date': builty[3],
        'rake_point_name': builty[4],
        'account_id': builty[5],
        'warehouse_id': builty[6],
        'cgmf_id': builty[7],
        'truck_id': builty[8],
        'loading_point': builty[9],
        'unloading_point': builty[10],
        'goods_name': builty[11],
        'number_of_bags': builty[12],
        'quantity_mt': builty[13],
        'kg_per_bag': builty[14],
        'rate_per_mt': builty[15],
        'total_freight': builty[16],
        'advance': builty[17],
        'to_pay': builty[18],
        'lr_number': builty[19],
        'lr_index': builty[20],
        'created_by_role': builty[21],
        'created_at': builty[22],
        'account_name': builty[23],
        'warehouse_name': builty[24],
        'truck_number': builty[25],
        'driver_name': builty[26],
        'driver_mobile': builty[27],
        'owner_name': builty[28],
        'owner_mobile': builty[29],
        'builty_head': builty[30] if len(builty) > 30 else None,
        'receiver_name': builty[31] if len(builty) > 31 else None,
        'received_quantity': builty[32] if len(builty) > 32 else None,
        'account_address': builty[33] if len(builty) > 33 else None,
        'company_name': builty[34] if len(builty) > 34 else None,
        'company_code': builty[35] if len(builty) > 35 else None,
        'account_type': builty[36] if len(builty) > 36 else None,
        'sub_head': builty[37] if len(builty) > 37 else None,
        'cgmf_society_name': None
    }
    # Lookup CGMF society name if this builty is for a CGMF account
    if builty_dict['cgmf_id']:
        cgmf_row = db.execute_custom_query('SELECT society_name FROM cgmf WHERE id = ?', (builty_dict['cgmf_id'],))
        if cgmf_row:
            builty_dict['cgmf_society_name'] = cgmf_row[0][0]
    # Override account_type to 'Payal' if account name contains 'payal'
    if builty_dict['account_name'] and 'payal' in str(builty_dict['account_name']).lower():
        builty_dict['account_type'] = 'Payal'
    
    return render_template('print_builty.html', builty=builty_dict)

# ========== RAKE POINT Dashboard & Routes ==========

@app.route('/rakepoint/dashboard')
@login_required
def rakepoint_dashboard():
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get all data
    all_builties = db.get_all_builties()
    all_loading_slips = db.get_all_loading_slips_with_status()
    
    # Recent builties (last 10)
    recent_builties = all_builties[:10]
    
    # Calculate statistics
    total_builties = len(all_builties)
    total_loading_slips = len(all_loading_slips)
    
    # Use optimized single-query method instead of loop
    active_rakes = db.count_active_rakes_optimized()
    
    # Today's builties - PostgreSQL returns date objects, SQLite returns strings
    today = datetime.now().date()
    today_builties = sum(1 for builty in all_builties 
                        if builty[3] and (builty[3] == today or 
                           (isinstance(builty[3], str) and builty[3][:10] == str(today))))
    
    # Use optimized method for rakes with balances
    all_rakes = db.get_rakes_with_balances()
    
    return render_template('rakepoint/dashboard.html', 
                         recent_builties=recent_builties,
                         rakes=all_rakes,
                         total_builties=total_builties,
                         total_loading_slips=total_loading_slips,
                         active_rakes=active_rakes,
                         today_builties=today_builties)

@app.route('/rakepoint/create-builty', methods=['GET', 'POST'])
@login_required
def rakepoint_create_builty():
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Generate builty number (can be auto-generated based on date + sequence)
        builty_number = f"BLT-{datetime.now().strftime('%Y%m%d')}-{datetime.now().strftime('%H%M%S')}"
        
        # Get loading_slip_id to link builty to loading slip
        loading_slip_id = request.form.get('loading_slip_id')
        
        # Get 14 fields from form + rake_code
        rake_code = request.form.get('rake_code')  # NEW: Link to rake
        date = request.form.get('date') or request.form.get('builty_date')  # Support both field names
        rake_point_name = request.form.get('rake_point_name')
        account_warehouse = request.form.get('account_warehouse')  # Field 3
        truck_number = request.form.get('truck_number')  # Field 4
        loading_point = request.form.get('loading_point')  # Field 5
        unloading_point = request.form.get('unloading_point')  # Field 6
        truck_driver = request.form.get('truck_driver')  # Field 7
        truck_owner = request.form.get('truck_owner')  # Field 8
        mobile_number_1 = request.form.get('mobile_number_1')  # Field 9
        mobile_number_2 = request.form.get('mobile_number_2')  # Field 10
        goods_name = request.form.get('goods_name')  # Field 11
        number_of_bags = int(request.form.get('number_of_bags'))  # Field 12
        quantity_wt_mt = float(request.form.get('quantity_wt_mt'))  # Field 13
        freight_details = request.form.get('freight_details')  # Field 14
        lr_number = request.form.get('lr_number')  # Field 15
        
        # Determine if account_warehouse is account, warehouse, or CGMF
        account_id = None
        warehouse_id = None
        cgmf_id = None
        
        # FIRST: Try to get IDs directly from loading slip (most reliable)
        slip_account_id = request.form.get('slip_account_id')
        slip_warehouse_id = request.form.get('slip_warehouse_id')
        slip_cgmf_id = request.form.get('slip_cgmf_id')
        
        if slip_account_id:
            account_id = slip_account_id
            print(f"DEBUG: Using account_id {account_id} from loading slip")
        elif slip_warehouse_id:
            warehouse_id = slip_warehouse_id
            print(f"DEBUG: Using warehouse_id {warehouse_id} from loading slip")
        elif slip_cgmf_id:
            cgmf_id = slip_cgmf_id
            print(f"DEBUG: Using cgmf_id {cgmf_id} from loading slip")
        # FALLBACK: Check if it's a CGMF account by name prefix
        elif account_warehouse and account_warehouse.startswith('CGMF:'):
            cgmf_id = account_warehouse.split(':')[1]
            print(f"DEBUG: Using cgmf_id {cgmf_id} from CGMF prefix")
        elif account_warehouse:
            # FALLBACK: Simple check for account or warehouse by name
            accounts = db.get_all_accounts()
            warehouses = db.get_all_warehouses()
            
            for account in accounts:
                if account[1] == account_warehouse:
                    account_id = account[0]
                    print(f"DEBUG: Found account_id {account_id} by name match")
                    break
            
            if account_id is None:
                for warehouse in warehouses:
                    if warehouse[1] == account_warehouse:
                        warehouse_id = warehouse[0]
                        print(f"DEBUG: Found warehouse_id {warehouse_id} by name match")
                        break
        
        # Check if truck exists, if not create it
        truck = db.get_truck_by_number(truck_number)
        if not truck:
            truck_id = db.add_truck(truck_number, truck_driver, mobile_number_1, truck_owner, mobile_number_2)
        else:
            truck_id = truck[0]
        
        # Parse freight details - assuming it's a number or contains a number
        try:
            total_freight = float(freight_details.replace('₹', '').replace(',', '').strip())
        except:
            total_freight = 0.0
        
        # Parse freight details and advance
        freight_advance = float(request.form.get('freight_advance', 0))
        to_pay = float(request.form.get('to_pay', 0))
        
        # Calculate kg per bag (assuming 50kg per bag as default)
        kg_per_bag = (quantity_wt_mt * 1000) / number_of_bags if number_of_bags > 0 else 50
        rate_per_mt = total_freight / quantity_wt_mt if quantity_wt_mt > 0 else 0
        
        # Get sub_head and receiver details
        sub_head = request.form.get('sub_head', '')
        receiver_name = request.form.get('receiver_name', '')
        received_quantity = request.form.get('received_quantity')
        received_quantity = float(received_quantity) if received_quantity else None
        supply_term = request.form.get('supply_term', 'EX')
        
        builty_id = db.add_builty(builty_number, rake_code, date, rake_point_name, account_id, warehouse_id,
                                  truck_id, loading_point, unloading_point, goods_name, number_of_bags,
                                  quantity_wt_mt, kg_per_bag, rate_per_mt, total_freight, freight_advance, to_pay, lr_number,
                                  0, 'rakepoint', cgmf_id, sub_head, receiver_name, received_quantity, supply_term)
        
        if builty_id:
            # Link the loading slip to this builty
            if loading_slip_id:
                link_success = db.link_loading_slip_to_builty(loading_slip_id, builty_id)
                if link_success:
                    flash(f'Builty {builty_number} created and linked to loading slip successfully!', 'success')
                else:
                    flash(f'Builty {builty_number} created but failed to link loading slip. Please contact admin.', 'warning')
            else:
                flash(f'Builty {builty_number} created successfully!', 'success')
            return redirect(url_for('rakepoint_dashboard'))
        else:
            flash('Error creating builty. Please try again.', 'error')
    
    accounts = db.get_all_accounts()
    warehouses = db.get_all_warehouses()
    rakes = db.get_all_rakes()  # NEW: Get all rakes for selection
    loading_slips = db.get_rakepoint_loading_slips()  # Get only rake loading slips (exclude warehouse ones)
    cgmf_list = db.get_all_cgmf()
    
    return render_template('rakepoint/create_builty.html', 
                         accounts=accounts, 
                         warehouses=warehouses,
                         rakes=rakes,
                         loading_slips=loading_slips,
                         cgmf_list=cgmf_list)

@app.route('/rakepoint/create-loading-slip', methods=['GET', 'POST'])
@login_required
def rakepoint_create_loading_slip():
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            rake_code = request.form.get('rake_code')
            loading_point = request.form.get('loading_point')
            destination = request.form.get('destination')
            slip_date = request.form.get('slip_date') or None

            # SERVER-SIDE serial number generation to prevent duplicates on quick consecutive submits
            serial_result = db.execute_custom_query(
                'SELECT COALESCE(MAX(CAST(slip_number AS INTEGER)), 0) + 1 FROM loading_slips WHERE rake_code = %s',
                (rake_code,)
            )
            serial_number = serial_result[0][0] if serial_result else 1

            # DUPLICATE CHECK: Verify this slip doesn't already exist
            existing_slip = db.execute_custom_query('''
                SELECT id FROM loading_slips 
                WHERE rake_code = %s AND slip_number = %s
            ''', (rake_code, serial_number))
            
            if existing_slip:
                flash(f'Loading slip #{serial_number} for rake {rake_code} already exists!', 'warning')
                return redirect(url_for('rakepoint_create_loading_slip'))
            
            account = request.form.get('account')

            # Support multi-product (product_name[] array) or single-product (goods_name)
            product_names = request.form.getlist('product_name[]')
            use_multi_product = bool(product_names and any(p for p in product_names))

            if use_multi_product:
                qty_bags_list = request.form.getlist('quantity_bags[]')
                qty_mt_list   = request.form.getlist('quantity_mt[]')
                products_data = []
                total_bags = 0
                total_mt   = 0.0
                for pname, bags, mt in zip(product_names, qty_bags_list, qty_mt_list):
                    if not pname:
                        continue
                    b = int(bags) if bags else 0
                    m = float(mt) if mt else 0.0
                    total_bags += b
                    total_mt   += m
                    products_data.append({'product_name': pname, 'quantity_bags': b, 'quantity_mt': m})
                goods_name       = products_data[0]['product_name'] if products_data else ''
                quantity_in_bags = total_bags
                quantity_in_mt   = total_mt
            else:
                goods_name       = request.form.get('goods_name')
                quantity_in_bags = int(request.form.get('quantity_in_bags', 0))
                quantity_in_mt   = float(request.form.get('quantity_in_mt', 0))
                products_data    = None
            truck_number = request.form.get('truck_number')
            wagon_number = request.form.get('wagon_number', '')
            truck_driver = request.form.get('truck_driver')
            truck_owner = request.form.get('truck_owner')
            mobile_number_1 = request.form.get('mobile_number_1')
            mobile_number_2 = request.form.get('mobile_number_2', '')
            truck_details = request.form.get('truck_details', '')
            sub_head = request.form.get('sub_head', '')  # Sub head for Payal accounts
            warehouse_account_type = request.form.get('warehouse_account_type', '')  # Account type for warehouse stock
            warehouse_account_id = request.form.get('warehouse_account_id', '')  # Account ID for warehouse stock
            
            # Convert warehouse_account_id to int if provided
            warehouse_account_id = warehouse_account_id if warehouse_account_id else None
            
            # CRITICAL: Check rake quantity balance before creating loading slip
            rake_balance = db.get_rake_balance(rake_code)
            if not rake_balance:
                flash('Error: Invalid rake code', 'error')
                return redirect(url_for('rakepoint_create_loading_slip'))

            # Block slips for closed rakes (safety check even though UI hides them)
            rake_status = db.execute_custom_query(
                'SELECT is_closed FROM rakes WHERE rake_code = %s', (rake_code,)
            )
            if rake_status and rake_status[0][0]:
                flash('This rake is closed. No more loading slips can be created.', 'error')
                return redirect(url_for('rakepoint_create_loading_slip'))

            if quantity_in_mt > rake_balance['remaining']:
                flash(f'Error: Insufficient rake quantity! Remaining: {rake_balance["remaining"]:.2f} MT, Requested: {quantity_in_mt:.2f} MT', 'error')
                return redirect(url_for('rakepoint_create_loading_slip'))
            
            # Determine if dispatch is to account, warehouse, or CGMF
            accounts = db.get_all_accounts()
            warehouses = db.get_all_warehouses()
            cgmf_list = db.get_all_cgmf()
            account_id = None
            warehouse_id = None
            cgmf_id = None
            
            # Check if it's a CGMF (format: CGMF:<id>)
            if account and account.startswith('CGMF:'):
                cgmf_id = account.split(':')[1]
                print(f"DEBUG: Loading slip dispatch to CGMF {cgmf_id}")
            else:
                # Check if it's a warehouse FIRST (warehouses are more specific)
                for wh in warehouses:
                    if wh[1] == account:
                        warehouse_id = wh[0]
                        print(f"DEBUG: Loading slip dispatch to WAREHOUSE '{account}' (id={warehouse_id})")
                        break
                
                # If not found in warehouses, check accounts
                if warehouse_id is None:
                    for acc in accounts:
                        if acc[1] == account:
                            account_id = acc[0]
                            print(f"DEBUG: Loading slip dispatch to ACCOUNT '{account}' (id={account_id})")
                            break
                
                if warehouse_id is None and account_id is None:
                    print(f"DEBUG: Could not find '{account}' in warehouses or accounts!")
            
            # Check if truck exists, if not create it with driver/owner details
            truck = db.get_truck_by_number(truck_number)
            if not truck:
                truck_id = db.add_truck(truck_number, truck_driver, mobile_number_1, truck_owner, mobile_number_2)
            else:
                truck_id = truck[0]
            
            slip_id = db.add_loading_slip(rake_code, serial_number, loading_point, destination,
                                          account_id, warehouse_id, quantity_in_bags, quantity_in_mt, truck_id, 
                                          wagon_number, goods_name, truck_driver, truck_owner,
                                          mobile_number_1, mobile_number_2, truck_details, None, cgmf_id, sub_head,
                                          warehouse_account_id, warehouse_account_type, slip_date)
            
            # slip_id is a UUID string (success) or None (failure)
            if slip_id is not None:
                # Save individual product rows for multi-product slips
                if use_multi_product and products_data:
                    db.add_loading_slip_products(slip_id, products_data)

                # CRITICAL: Invalidate cache after successful write to prevent stale data
                db.invalidate_cache()
                
                flash(f'Loading slip #{serial_number} created successfully!', 'success')
                if request.form.get('action') == 'print':
                    # Use redirect with print flag to prevent form resubmission on refresh
                    return redirect(url_for('rakepoint_create_loading_slip', print_slip=slip_id))
                return redirect(url_for('rakepoint_dashboard'))
            else:
                flash('Error creating loading slip. Please try again.', 'error')
        
        except ValueError as e:
            flash(f'Invalid input data: {str(e)}', 'error')
        except Exception as e:
            print(f"Error in create_loading_slip: {e}")
            import traceback
            traceback.print_exc()
            flash('An error occurred while creating the loading slip. Please try again.', 'error')
    
    rakes = db.get_active_rakes()
    accounts = db.get_all_accounts()
    warehouses = db.get_all_warehouses()
    companies = db.get_all_companies()
    cgmf_list = db.get_all_cgmf()
    trucks = db.get_all_trucks()
    builties = db.get_all_builties()
    
    # Check if we need to print a slip (from POST redirect)
    print_slip_id = request.args.get('print_slip')
    
    return render_template('rakepoint/create_loading_slip.html', 
                         rakes=rakes,
                         accounts=accounts,
                         warehouses=warehouses,
                         companies=companies,
                         cgmf_list=cgmf_list,
                         trucks=trucks,
                         builties=builties,
                         print_slip_id=print_slip_id)

@app.route('/api/rake-balance/<path:rake_code>')
@login_required
def get_rake_balance_api(rake_code):
    """API endpoint to get rake balance"""
    if current_user.role != 'RakePoint':
        return jsonify({'error': 'Unauthorized'}), 403
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    balance = db.get_rake_balance(rake_code)
    if balance:
        return jsonify(balance)
    else:
        return jsonify({'error': 'Rake not found'}), 404

@app.route('/api/rake-products/<path:rake_code>')
@login_required
def get_rake_products_api(rake_code):
    """API endpoint to get products for a rake"""
    if current_user.role not in ['RakePoint', 'Warehouse', 'Admin']:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    products = db.get_rake_products(rake_code)
    print(f"[DEBUG] Rake products for {rake_code}: {products}")
    if products:
        result = {
            'products': [
                {
                    'rake_product_id': p[0],
                    'product_id': p[1],
                    'product_name': p[2] or '',
                    'product_code': p[3] or '',
                    'quantity_mt': float(p[4]) if p[4] else 0,
                    'unit_per_bag': float(p[5]) if p[5] else 50
                }
                for p in products
            ]
        }
        print(f"[DEBUG] API response: {result}")
        return jsonify(result)
    else:
        # Fallback: If no rake_products, get from rakes table (backward compatibility)
        rake = db.execute_custom_query('''
            SELECT product_name, product_code, rr_quantity
            FROM rakes WHERE rake_code = ?
        ''', (rake_code,))
        if rake:
            return jsonify({
                'products': [{
                    'rake_product_id': 0,
                    'product_id': 0,
                    'product_name': rake[0][0],
                    'product_code': rake[0][1],
                    'quantity_mt': rake[0][2],
                    'unit_per_bag': 50
                }]
            })
        return jsonify({'products': []})

@app.route('/api/next-serial-number/<path:rake_code>')
@login_required
def get_next_serial_number_api(rake_code):
    """API endpoint to get next serial number for rake"""
    if current_user.role != 'RakePoint':
        return jsonify({'error': 'Unauthorized'}), 403
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    next_serial = db.get_next_serial_number_for_rake(rake_code)
    return jsonify({'next_serial': next_serial})

@app.route('/api/next-lr-number')
@login_required
def get_next_lr_number_api():
    """API endpoint to get next LR number"""
    if current_user.role not in ['RakePoint', 'Warehouse']:
        return jsonify({'error': 'Unauthorized'}), 403
    
    next_lr = db.get_next_lr_number()
    return jsonify({'next_lr': next_lr})

@app.route('/api/next-warehouse-serial/<string:warehouse_id>')
@login_required
def get_next_warehouse_serial_api(warehouse_id):
    """API endpoint to get next warehouse stock serial number"""
    if current_user.role not in ['Warehouse', 'Admin']:
        return jsonify({'error': 'Unauthorized'}), 403
    
    next_serial = db.get_next_warehouse_stock_serial(warehouse_id)
    return jsonify({'serial_number': next_serial})

@app.route('/api/builty-details/<string:builty_id>')
@login_required
def get_builty_details_api(builty_id):
    """API endpoint to get builty details for auto-filling warehouse stock in form"""
    if current_user.role not in ['Warehouse', 'Admin']:
        return jsonify({'error': 'Unauthorized'}), 403
    
    builty = db.get_builty_by_id(builty_id)
    
    if not builty:
        return jsonify({'error': 'Builty not found'}), 404
    
    # Get rake details to find company and product info
    rake = db.execute_custom_query('''
        SELECT r.company_name, r.product_name, r.rake_point_name
        FROM rakes r
        WHERE r.rake_code = ?
    ''', (builty[2],))
    
    company_name = rake[0][0] if rake else None
    product_name = rake[0][1] if rake else builty[11]  # fallback to goods_name
    
    # Find company_id by company_name
    company_id = None
    if company_name:
        company = db.execute_custom_query('SELECT id FROM companies WHERE company_name = ?', (company_name,))
        if company:
            company_id = company[0][0]
    
    # Find product_id by product_name
    product_id = None
    if product_name:
        product = db.execute_custom_query('SELECT id FROM products WHERE product_name = ?', (product_name,))
        if product:
            product_id = product[0][0]
    
    return jsonify({
        'builty_number': builty[1],
        'goods_name': builty[11],
        'quantity_mt': float(builty[13]) if builty[13] else 0,
        'company_id': company_id,
        'product_id': product_id,
        'account_id': builty[5],
        'warehouse_id': builty[6]
    })

@app.route('/api/next-loading-slip-serial/<warehouse_name>')
@login_required
def get_next_loading_slip_serial_api(warehouse_name):
    """API endpoint to get next loading slip serial number for warehouse"""
    if current_user.role not in ['Warehouse', 'Admin']:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get the maximum serial number for this warehouse
    result = db.execute_custom_query('''
        SELECT MAX(CAST(slip_number AS INTEGER)) 
        FROM loading_slips 
        WHERE loading_point_name = ?
    ''', (warehouse_name,))
    
    max_serial = result[0][0] if result and result[0][0] else 0
    next_serial = max_serial + 1
    
    return jsonify({'serial_number': next_serial})


@app.route('/api/warehouse-products/<warehouse_id>')
@login_required
def get_warehouse_products_api(warehouse_id):
    """API endpoint to get products with positive stock balance in a warehouse."""
    if current_user.role not in ['Warehouse', 'Admin']:
        return jsonify({'error': 'Unauthorized'}), 403
    products = db.get_products_in_warehouse(warehouse_id)
    return jsonify({'products': products})

@app.route('/api/account-dispatches/<account_name>')
@login_required
def get_account_dispatches(account_name):
    """API endpoint to get all dispatches for a specific account"""
    if current_user.role != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    dispatches = db.execute_custom_query('''
        SELECT b.rake_code, b.builty_number, b.date, b.quantity_mt, b.goods_name
        FROM builty b
        JOIN accounts a ON b.account_id = a.id
        WHERE a.account_name = ?
        ORDER BY b.date DESC, b.created_at DESC
    ''', (account_name,))
    
    result = []
    if dispatches:
        for d in dispatches:
            result.append({
                'rake_code': d[0] or 'N/A',
                'builty_number': d[1],
                'date': d[2],
                'quantity_mt': d[3],
                'goods_name': d[4] or 'N/A'
            })
    
    return jsonify(result)

@app.route('/api/cgmf-dispatches/<society_name>')
@login_required
def get_cgmf_dispatches(society_name):
    """API endpoint to get all dispatches for a specific CGMF society"""
    if current_user.role != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    dispatches = db.execute_custom_query('''
        SELECT b.rake_code, b.builty_number, b.date, b.quantity_mt, b.goods_name
        FROM builty b
        JOIN cgmf c ON b.cgmf_id = c.id
        WHERE c.society_name = ?
        ORDER BY b.date DESC, b.created_at DESC
    ''', (society_name,))
    
    result = []
    if dispatches:
        for d in dispatches:
            result.append({
                'rake_code': d[0] or 'N/A',
                'builty_number': d[1],
                'date': d[2],
                'quantity_mt': d[3],
                'goods_name': d[4] or 'N/A'
            })
    
    return jsonify(result)

@app.route('/rakepoint/loading-slips')
@login_required
def rakepoint_loading_slips():
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    loading_slips = db.get_all_loading_slips_with_status()
    
    return render_template('rakepoint/loading_slips.html', loading_slips=loading_slips)

@app.route('/rakepoint/print-loading-slip/<string:slip_id>')
@login_required
def rakepoint_print_loading_slip(slip_id):
    """Print a specific loading slip"""
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get loading slip details
    slip = db.execute_custom_query('''
        SELECT ls.id, ls.rake_code, ls.slip_number, ls.loading_point_name,
               ls.destination, ls.quantity_bags, ls.quantity_mt, ls.goods_name,
               ls.truck_driver, ls.truck_owner, ls.mobile_number_1, ls.mobile_number_2,
               ls.wagon_number, ls.created_at, t.truck_number,
               COALESCE(a.account_name, w.warehouse_name, cg.society_name) as destination_name,
               COALESCE(a.account_type::text, CASE WHEN cg.id IS NOT NULL THEN 'CGMF' ELSE 'Warehouse' END) as destination_type
        FROM loading_slips ls
        LEFT JOIN trucks t ON ls.truck_id = t.id
        LEFT JOIN accounts a ON ls.account_id = a.id
        LEFT JOIN warehouses w ON ls.warehouse_id = w.id
        LEFT JOIN cgmf cg ON ls.cgmf_id = cg.id
        WHERE ls.id = ?
    ''', (slip_id,))
    
    if not slip:
        flash('Loading slip not found', 'error')
        return redirect(url_for('rakepoint_loading_slips'))
    
    return render_template('print_loading_slip.html', slip=slip[0])

@app.route('/rakepoint/print-builty/<string:builty_id>')
@login_required
def rakepoint_print_builty(builty_id):
    """RakePoint can print builty"""
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get builty details with all related information
    builty = db.get_builty_by_id(builty_id)
    
    if not builty:
        flash('Builty not found', 'error')
        return redirect(url_for('rakepoint_dashboard'))
    
    # Convert to dict for easier template access
    # Explicit column order from get_builty_by_id query
    builty_dict = {
        'builty_id': builty[0],
        'builty_number': builty[1],
        'rake_code': builty[2],
        'date': builty[3],
        'rake_point_name': builty[4],
        'account_id': builty[5],
        'warehouse_id': builty[6],
        'cgmf_id': builty[7],
        'truck_id': builty[8],
        'loading_point': builty[9],
        'unloading_point': builty[10],
        'goods_name': builty[11],
        'number_of_bags': builty[12],
        'quantity_mt': builty[13],
        'kg_per_bag': builty[14],
        'rate_per_mt': builty[15],
        'total_freight': builty[16],
        'advance': builty[17],
        'to_pay': builty[18],
        'lr_number': builty[19],
        'lr_index': builty[20],
        'created_by_role': builty[21],
        'created_at': builty[22],
        'account_name': builty[23],
        'warehouse_name': builty[24],
        'truck_number': builty[25],
        'driver_name': builty[26],
        'driver_mobile': builty[27],
        'owner_name': builty[28],
        'owner_mobile': builty[29],
        'builty_head': builty[30] if len(builty) > 30 else None,
        'receiver_name': builty[31] if len(builty) > 31 else None,
        'received_quantity': builty[32] if len(builty) > 32 else None,
        'account_address': builty[33] if len(builty) > 33 else None,
        'company_name': builty[34] if len(builty) > 34 else None,
        'company_code': builty[35] if len(builty) > 35 else None,
        'account_type': builty[36] if len(builty) > 36 else None,
        'sub_head': builty[37] if len(builty) > 37 else None,
        'cgmf_society_name': None
    }
    # Lookup CGMF society name if this builty is for a CGMF account
    if builty_dict['cgmf_id']:
        cgmf_row = db.execute_custom_query('SELECT society_name FROM cgmf WHERE id = ?', (builty_dict['cgmf_id'],))
        if cgmf_row:
            builty_dict['cgmf_society_name'] = cgmf_row[0][0]
    # Override account_type to 'Payal' if account name contains 'payal'
    if builty_dict['account_name'] and 'payal' in str(builty_dict['account_name']).lower():
        builty_dict['account_type'] = 'Payal'
    
    return render_template('print_builty.html', builty=builty_dict)

@app.route('/rakepoint/loading-slips/<path:rake_code>')
@login_required
def rakepoint_view_loading_slips(rake_code):
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # URL decode the rake_code to handle special characters like &, spaces, etc.
    rake_code = unquote(rake_code)
    
    rake = db.get_rake_by_code(rake_code)
    slips = db.get_loading_slips_by_rake(rake_code)
    
    return render_template('rakepoint/loading_slips.html', rake=rake, loading_slips=slips)

@app.route('/rakepoint/all-builties')
@login_required
def rakepoint_all_builties():
    """View all builties created by rakepoint"""
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get all builties created by rakepoint
    builties = db.execute_custom_query('''
        SELECT b.id, b.builty_number, b.rake_code, b.date, b.rake_point_name,
               b.account_id, b.warehouse_id, b.cgmf_id, b.truck_id,
               b.loading_point, b.unloading_point, b.goods_name, b.number_of_bags,
               b.quantity_mt, b.kg_per_bag, b.rate_per_mt, b.total_freight,
               b.advance, b.to_pay, b.lr_number, b.lr_index, b.created_by_role,
               b.sub_head, b.receiver_name, b.received_quantity, b.created_at,
               a.account_name, w.warehouse_name, t.truck_number
        FROM builty b
        LEFT JOIN accounts a ON b.account_id = a.id
        LEFT JOIN warehouses w ON b.warehouse_id = w.id
        LEFT JOIN trucks t ON b.truck_id = t.id
        WHERE b.created_by_role = 'rakepoint'
        ORDER BY b.created_at DESC
    ''')
    
    return render_template('rakepoint/all_builties.html', builties=builties)

@app.route('/rakepoint/view-ebills')
@login_required
def rakepoint_view_ebills():
    """RakePoint can view e-bills for builties they created"""
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    ebills = db.get_ebills_by_builty_creator('rakepoint')
    return render_template('rakepoint/view_ebills.html', ebills=ebills)

@app.route('/rakepoint/download-bill/<filename>')
@login_required
def rakepoint_download_bill(filename):
    """RakePoint can download bill PDFs for their builties"""
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('rakepoint_view_ebills'))

@app.route('/rakepoint/download-eway-bill/<filename>')
@login_required
def rakepoint_download_eway_bill(filename):
    """RakePoint can download eway bill PDFs for their builties"""
    if current_user.role != 'RakePoint':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'eway_bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('rakepoint_view_ebills'))

# ========== WAREHOUSE Dashboard & Routes ==========

@app.route('/warehouse/dashboard')
@login_required
def warehouse_dashboard():
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    warehouses = db.get_all_warehouses()
    
    # Use batch query instead of loop - single query for all warehouse balances
    all_balances = db.get_all_warehouse_balances()
    
    warehouse_stats = []
    total_stock_in = 0
    total_stock_out = 0
    
    for warehouse in warehouses:
        wh_id = warehouse[0]
        balance = all_balances.get(wh_id, {'stock_in': 0, 'stock_out': 0, 'balance': 0})
        warehouse_stats.append({
            'warehouse': warehouse,
            'stock_in': balance['stock_in'],
            'stock_out': balance['stock_out'],
            'balance': balance['balance']
        })
        total_stock_in += balance['stock_in']
        total_stock_out += balance['stock_out']
    
    # Use optimized method for recent movements
    recent_movements = db.get_recent_warehouse_movements(limit=10)
    
    # Use optimized method for dashboard stats
    stats = db.get_warehouse_dashboard_stats()
    
    # Support both legacy schema (*_id PKs) and newer schema (id PKs).
    def _pk_col(table_name, legacy_pk):
        try:
            # SQLite / libsql path
            cols = db.execute_custom_query(
                f"SELECT name FROM pragma_table_info('{table_name}')"
            ) or []
            names = [c[0] for c in cols if c and len(c) > 0]
            if names:
                return 'id' if 'id' in names else legacy_pk
        except Exception:
            pass

        try:
            # PostgreSQL path
            cols = db.execute_custom_query(
                f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'"
            ) or []
            names = [c[0] for c in cols if c and len(c) > 0]
            if names:
                return 'id' if 'id' in names else legacy_pk
        except Exception:
            pass

        return legacy_pk

    account_pk = _pk_col('accounts', 'account_id')
    cgmf_pk = _pk_col('cgmf', 'cgmf_id')
    company_pk = _pk_col('companies', 'company_id')
    product_pk = _pk_col('products', 'product_id')
    warehouse_pk = _pk_col('warehouses', 'warehouse_id')

    # Get stock by Account, CGMF, and Company for the dashboard
    account_stock = db.execute_custom_query(f'''
        SELECT w.{warehouse_pk}, w.warehouse_name, a.{account_pk}, a.account_name,
               CASE
                   WHEN LOWER(COALESCE(a.account_type, '')) IN ('company', 'payal') THEN 'Payal'
                   WHEN LOWER(COALESCE(a.account_type, '')) = 'dealer' THEN 'Dealer'
                   WHEN LOWER(COALESCE(a.account_type, '')) = 'retailer' THEN 'Retailer'
                   ELSE COALESCE(a.account_type, 'Account')
               END AS account_type,
               COALESCE(p.product_name, 'N/A') AS product_name,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.{warehouse_pk}
        JOIN accounts a ON ws.account_id = a.{account_pk}
        LEFT JOIN products p ON ws.product_id = p.{product_pk}
        WHERE ws.account_id IS NOT NULL
        GROUP BY w.{warehouse_pk}, w.warehouse_name, a.{account_pk}, a.account_name,
                 CASE
                     WHEN LOWER(COALESCE(a.account_type, '')) IN ('company', 'payal') THEN 'Payal'
                     WHEN LOWER(COALESCE(a.account_type, '')) = 'dealer' THEN 'Dealer'
                     WHEN LOWER(COALESCE(a.account_type, '')) = 'retailer' THEN 'Retailer'
                     ELSE COALESCE(a.account_type, 'Account')
                 END,
                 COALESCE(p.product_name, 'N/A')
        HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ORDER BY w.warehouse_name, a.account_name, COALESCE(p.product_name, 'N/A')
    ''') or []
    
    cgmf_stock = db.execute_custom_query(f'''
        SELECT w.{warehouse_pk}, w.warehouse_name, c.{cgmf_pk}, c.society_name, 'CGMF' as type,
               COALESCE(p.product_name, 'N/A') AS product_name,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.{warehouse_pk}
        JOIN cgmf c ON ws.cgmf_id = c.{cgmf_pk}
        LEFT JOIN products p ON ws.product_id = p.{product_pk}
        WHERE ws.cgmf_id IS NOT NULL
        GROUP BY w.{warehouse_pk}, w.warehouse_name, c.{cgmf_pk}, c.society_name, COALESCE(p.product_name, 'N/A')
        HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ORDER BY w.warehouse_name, c.society_name, COALESCE(p.product_name, 'N/A')
    ''') or []
    
    company_stock = db.execute_custom_query(f'''
        SELECT w.{warehouse_pk}, w.warehouse_name, c.{company_pk}, c.company_name, 'Company' as type,
               COALESCE(p.product_name, 'N/A') AS product_name,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.{warehouse_pk}
        JOIN companies c ON ws.company_id = c.{company_pk}
        LEFT JOIN products p ON ws.product_id = p.{product_pk}
        WHERE ws.company_id IS NOT NULL AND ws.account_id IS NULL AND ws.cgmf_id IS NULL
        GROUP BY w.{warehouse_pk}, w.warehouse_name, c.{company_pk}, c.company_name, COALESCE(p.product_name, 'N/A')
        HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ORDER BY w.warehouse_name, c.company_name, COALESCE(p.product_name, 'N/A')
    ''') or []
    
    all_stock = account_stock + cgmf_stock + company_stock
    
    return render_template('warehouse/dashboard.html', 
                         warehouse_stats=warehouse_stats,
                         total_stock_in=total_stock_in,
                         total_stock_out=total_stock_out,
                         warehouse_count=len(warehouses),
                         recent_movements=recent_movements,
                         account_count=stats['account_count'],
                         builty_count=stats['builty_count'],
                         today_stock_in=stats['today_stock_in'],
                         account_stock=all_stock)

@app.route('/warehouse/stock-in', methods=['GET', 'POST'])
@login_required
def warehouse_stock_in():
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        serial_number = request.form.get('serial_number')
        warehouse_id = request.form.get('warehouse_name')
        source_type = request.form.get('source_type', 'rake')
        builty_id = request.form.get('builty_number') if source_type == 'rake' else None
        company_id = request.form.get('company_id')
        product_id = request.form.get('product_id')
        quantity = float(request.form.get('unloaded_quantity'))
        employee_id = request.form.get('employee_id')

        # Resolve truck_id: look up by truck number (or create) to get the UUID FK
        truck_id = None
        if source_type == 'truck':
            truck_number_raw = (request.form.get('truck_number') or '').strip()
            if truck_number_raw:
                existing_truck = db.get_truck_by_number(truck_number_raw)
                if existing_truck:
                    truck_id = existing_truck[0]
                else:
                    truck_id = db.add_truck(truck_number_raw, None, None, None, None)
        
        # Handle account_id which may be a regular account, CGMF, or Company
        account_id_raw = request.form.get('account_id') or ''
        account_id = None
        cgmf_id = None
        source_company_id = None  # For when stock belongs to a specific company
        
        if account_id_raw.startswith('CGMF:'):
            cgmf_id = account_id_raw.replace('CGMF:', '')
        elif account_id_raw.startswith('COMPANY:'):
            source_company_id = account_id_raw.replace('COMPANY:', '')
        else:
            account_id = account_id_raw or None
        
        stock_in_date = request.form.get('stock_in_date')
        remarks = request.form.get('remarks', '')
        sub_head = request.form.get('sub_head', '')  # Sub head for Payal accounts
        
        # Convert builty_id and truck_id to integers if present
        if builty_id:
            pass  # builty_id is already a UUID string
            
            # Check if this builty has already been used for stock IN
            existing_stock = db.execute_custom_query('''
                SELECT ws.id, w.warehouse_name, ws.date, ws.quantity_mt
                FROM warehouse_stock ws
                JOIN warehouses w ON ws.warehouse_id = w.id
                WHERE ws.builty_id = ? AND ws.transaction_type = 'IN'
            ''', (builty_id,))
            
            if existing_stock:
                warehouse_name = existing_stock[0][1]
                stock_date = existing_stock[0][2]
                stock_quantity = existing_stock[0][3]
                flash(f'Error: This builty has already been recorded for Stock IN at {warehouse_name} on {stock_date} ({stock_quantity:.2f} MT). Cannot record duplicate stock IN for the same builty.', 'error')
                # Re-render form with all the data
                warehouses = db.get_all_warehouses()
                # Only show builties that were sent to warehouses (have warehouse_id) and not yet used for stock IN
                builties = db.execute_custom_query('''
                    SELECT b.*, w.warehouse_name
                    FROM builty b
                    JOIN warehouses w ON b.warehouse_id = w.id
                    LEFT JOIN warehouse_stock ws ON b.id = ws.builty_id AND ws.transaction_type = 'IN'
                    WHERE b.warehouse_id IS NOT NULL AND ws.id IS NULL
                    ORDER BY b.created_at DESC
                ''')
                trucks = db.get_all_trucks()
                companies = db.get_all_companies()
                products = db.get_all_products()
                employees = db.get_all_employees()
                accounts = db.get_all_accounts()
                cgmf_list = db.get_all_cgmf()
                recent_stock_in = db.execute_custom_query('''
                    SELECT ws.date, COALESCE(b.builty_number, 'Direct Truck') as builty_number, 
                           w.warehouse_name, COALESCE(a.account_name, 'N/A') as account_name,
                           ws.quantity_mt, COALESCE(e.employee_name, 'N/A') as employee_name
                    FROM warehouse_stock ws
                    JOIN warehouses w ON ws.warehouse_id = w.id
                    LEFT JOIN builty b ON ws.builty_id = b.id
                    LEFT JOIN accounts a ON ws.account_id = a.id
                    LEFT JOIN employees e ON ws.employee_id = e.id
                    WHERE ws.transaction_type = 'IN'
                    ORDER BY ws.date DESC, ws.created_at DESC
                    LIMIT 10
                ''')
                return render_template('warehouse/stock_in.html', 
                                     warehouses=warehouses,
                                     builties=builties,
                                     trucks=trucks,
                                     companies=companies,
                                     products=products,
                                     employees=employees,
                                     accounts=accounts,
                                     cgmf_list=cgmf_list,
                                     recent_stock_in=recent_stock_in)
        if truck_id:
            pass  # truck_id is already a UUID string
        
        # Map 'truck' to 'transfer' — the stock_source enum only has 'rake'/'transfer'
        db_source_type = 'transfer' if source_type == 'truck' else source_type

        stock_id = db.add_warehouse_stock_in(
            warehouse_id=warehouse_id,
            builty_id=builty_id,
            quantity_mt=quantity,
            employee_id=employee_id,
            account_id=account_id,
            cgmf_id=cgmf_id,
            date=stock_in_date,
            notes=remarks,
            company_id=company_id,
            product_id=product_id,
            source_type=db_source_type,
            truck_id=truck_id,
            serial_number=int(serial_number) if serial_number else None,
            sub_head=sub_head
        )
        
        if stock_id:
            flash(f'Stock IN recorded successfully! Serial No: {serial_number}', 'success')
            return redirect(url_for('warehouse_dashboard'))
        else:
            flash('Error recording stock IN', 'error')
    
    warehouses = db.get_all_warehouses()
    # Get only builties that were sent to warehouses (have warehouse_id) and haven't been used for stock IN yet
    builties = db.execute_custom_query('''
        SELECT b.*, w.warehouse_name as dest_warehouse_name
        FROM builty b
        JOIN warehouses w ON b.warehouse_id = w.id
        LEFT JOIN warehouse_stock ws ON b.id = ws.builty_id AND ws.transaction_type = 'IN'
        WHERE b.warehouse_id IS NOT NULL AND ws.id IS NULL
        ORDER BY b.created_at DESC
    ''')
    trucks = db.get_all_trucks()
    companies = db.get_all_companies()
    products = db.get_all_products()
    employees = db.get_all_employees()
    accounts = db.get_all_accounts()
    cgmf_list = db.get_all_cgmf()
    
    # Get recent stock IN entries for display
    recent_stock_in = db.execute_custom_query('''
        SELECT ws.date, COALESCE(b.builty_number, 'Direct Truck') as builty_number, 
               w.warehouse_name, COALESCE(a.account_name, 'N/A') as account_name,
               ws.quantity_mt, COALESCE(e.employee_name, 'N/A') as employee_name
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.id
        LEFT JOIN builty b ON ws.builty_id = b.id
        LEFT JOIN accounts a ON ws.account_id = a.id
        LEFT JOIN employees e ON ws.employee_id = e.id
        WHERE ws.transaction_type = 'IN'
        ORDER BY ws.date DESC, ws.created_at DESC
        LIMIT 10
    ''')
    
    return render_template('warehouse/stock_in.html', 
                         warehouses=warehouses,
                         builties=builties,
                         trucks=trucks,
                         companies=companies,
                         products=products,
                         employees=employees,
                         accounts=accounts,
                         cgmf_list=cgmf_list,
                         recent_stock_in=recent_stock_in)

@app.route('/warehouse/create-loading-slip', methods=['GET', 'POST'])
@login_required
def warehouse_create_loading_slip():
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            warehouse_name = request.form.get('warehouse_name')
            slip_date = request.form.get('slip_date') or None
            loading_point = request.form.get('loading_point')
            destination = request.form.get('destination')
            account = request.form.get('account')
            truck_number = request.form.get('truck_number')
            wagon_number = request.form.get('wagon_number', '')
            truck_driver = request.form.get('truck_driver')
            truck_owner = request.form.get('truck_owner')
            mobile_number_1 = request.form.get('mobile_number_1')
            mobile_number_2 = request.form.get('mobile_number_2', '')
            truck_details = request.form.get('truck_details', '')

            # Collect multi-product rows
            product_names = request.form.getlist('product_name[]')
            product_ids   = request.form.getlist('product_id[]')
            qty_bags_list = request.form.getlist('quantity_bags[]')
            qty_mt_list   = request.form.getlist('quantity_mt[]')

            if not product_names or all(not p for p in product_names):
                flash('At least one product is required.', 'error')
                return redirect(url_for('warehouse_create_loading_slip'))

            products_data = []
            total_bags = 0
            total_mt   = 0.0
            for pname, pid, bags, mt in zip(product_names, product_ids, qty_bags_list, qty_mt_list):
                if not pname:
                    continue
                b = int(bags) if bags else 0
                m = float(mt) if mt else 0.0
                total_bags += b
                total_mt   += m
                products_data.append({
                    'product_name': pname,
                    'product_id':   pid or None,
                    'quantity_bags': b,
                    'quantity_mt':   m
                })

            # Summary goods_name from first product (for loading_slips.goods_name)
            goods_name = products_data[0]['product_name'] if products_data else ''
            quantity_in_bags = total_bags
            quantity_in_mt   = total_mt

            # Get warehouse ID
            warehouse = db.execute_custom_query('SELECT id FROM warehouses WHERE warehouse_name = ?', (warehouse_name,))
            if not warehouse:
                flash('Invalid warehouse', 'error')
                return redirect(url_for('warehouse_create_loading_slip'))
            warehouse_id = warehouse[0][0]

            # SERVER-SIDE serial number generation to prevent duplicate on quick consecutive submits
            serial_result = db.execute_custom_query(
                'SELECT COALESCE(MAX(CAST(slip_number AS INTEGER)), 0) + 1 FROM loading_slips WHERE loading_point_name = ?',
                (warehouse_name,)
            )
            serial_number = serial_result[0][0] if serial_result else 1
            
            # Check if warehouse has enough stock
            balance = db.get_warehouse_balance_stock(warehouse_id)
            if balance['balance'] < quantity_in_mt:
                flash(f'Error: Warehouse has only {balance["balance"]:.2f} MT available. Cannot dispatch {quantity_in_mt:.2f} MT', 'error')
                return redirect(url_for('warehouse_create_loading_slip'))
            
            # Determine if dispatch is to account, warehouse, or CGMF
            accounts = db.get_all_accounts()
            warehouses_list = db.get_all_warehouses()
            cgmf_list = db.get_all_cgmf()
            account_id = None
            destination_warehouse_id = None
            cgmf_id = None
            
            # Check if it's a CGMF (format: CGMF:<id>)
            if account and account.startswith('CGMF:'):
                cgmf_id = account.split(':')[1]
            else:
                # Check if it's an account
                for acc in accounts:
                    if acc[1] == account:
                        account_id = acc[0]
                        break
                
                # If not found in accounts, check warehouses
                if account_id is None:
                    for wh in warehouses_list:
                        if wh[1] == account:
                            destination_warehouse_id = wh[0]
                            break
            
            # Check if truck exists, if not create it with driver/owner details
            truck = db.get_truck_by_number(truck_number)
            if not truck:
                truck_id = db.add_truck(truck_number, truck_driver, mobile_number_1, truck_owner, mobile_number_2)
            else:
                truck_id = truck[0]
            
            # Warehouse loading slips are secondary dispatches (stock already received
            # from rake). NULL rake_code prevents double-counting against rake totals.
            rake_code = 'WAREHOUSE'  # Must use sentinel; loading_slips.rake_code is NOT NULL

            # Add loading slip
            slip_id = db.add_loading_slip(rake_code, serial_number, loading_point, destination,
                                          account_id, destination_warehouse_id, quantity_in_bags, quantity_in_mt, 
                                          truck_id, wagon_number, goods_name, truck_driver, truck_owner,
                                          mobile_number_1, mobile_number_2, truck_details, None, cgmf_id,
                                          slip_date=slip_date)
            
            if slip_id:
                # Save individual product rows
                db.add_loading_slip_products(slip_id, products_data)

                # CRITICAL: Invalidate cache after successful write to prevent stale data
                db.invalidate_cache()
                
                flash(f'Loading slip #{serial_number} created successfully!', 'success')
                if request.form.get('action') == 'print':
                    # Use redirect to prevent form resubmission on refresh
                    return redirect(url_for('warehouse_create_loading_slip', print_slip=slip_id))
                return redirect(url_for('warehouse_dashboard'))
            else:
                flash('Error creating loading slip. Please try again.', 'error')
        
        except ValueError as e:
            flash(f'Invalid input data: {str(e)}', 'error')
        except Exception as e:
            print(f"Error in warehouse_create_loading_slip: {e}")
            import traceback
            traceback.print_exc()
            flash('An error occurred while creating the loading slip. Please try again.', 'error')
    
    warehouses = db.get_all_warehouses()
    accounts = db.get_all_accounts()
    cgmf_list = db.get_all_cgmf()
    trucks = db.get_all_trucks()
    products = db.get_all_products()
    
    # Check if we need to print a slip (from POST redirect)
    print_slip_id = request.args.get('print_slip')
    
    return render_template('warehouse/create_loading_slip.html',
                         warehouses=warehouses,
                         accounts=accounts,
                         cgmf_list=cgmf_list,
                         trucks=trucks,
                         products=products,
                         print_slip_id=print_slip_id)

@app.route('/warehouse/loading-slips')
@login_required
def warehouse_loading_slips():
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Show only loading slips created FROM warehouses (outbound dispatches)
    loading_slips = db.execute_custom_query('''
        SELECT ls.id, ls.rake_code, ls.slip_number, ls.loading_point_name, 
               ls.destination, 
               COALESCE(a.account_name, w.warehouse_name, cg.society_name) as destination_name,
               ls.quantity_bags, ls.quantity_mt, t.truck_number,
               ls.wagon_number, ls.builty_id, ls.created_at,
               ls.goods_name, ls.truck_driver, ls.truck_owner,
               ls.mobile_number_1, ls.mobile_number_2,
               b.builty_number
        FROM loading_slips ls
        LEFT JOIN accounts a ON ls.account_id = a.id
        LEFT JOIN warehouses w ON ls.warehouse_id = w.id
        LEFT JOIN cgmf cg ON ls.cgmf_id = cg.id
        LEFT JOIN trucks t ON ls.truck_id = t.id
        LEFT JOIN builty b ON ls.builty_id = b.id
        WHERE ls.loading_point_name IN (SELECT warehouse_name FROM warehouses)
        ORDER BY ls.created_at DESC
    ''')
    
    return render_template('warehouse/loading_slips.html', loading_slips=loading_slips)

@app.route('/warehouse/print-loading-slip/<string:slip_id>')
@login_required
def warehouse_print_loading_slip(slip_id):
    """Print a specific loading slip"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get loading slip details
    slip = db.execute_custom_query('''
        SELECT ls.id, ls.rake_code, ls.slip_number, ls.loading_point_name,
               ls.destination, ls.quantity_bags, ls.quantity_mt, ls.goods_name,
               ls.truck_driver, ls.truck_owner, ls.mobile_number_1, ls.mobile_number_2,
               ls.wagon_number, ls.created_at, t.truck_number,
               COALESCE(a.account_name, w.warehouse_name, cg.society_name) as destination_name,
               COALESCE(a.account_type::text, CASE WHEN cg.id IS NOT NULL THEN 'CGMF' ELSE 'Warehouse' END) as destination_type
        FROM loading_slips ls
        LEFT JOIN trucks t ON ls.truck_id = t.id
        LEFT JOIN accounts a ON ls.account_id = a.id
        LEFT JOIN warehouses w ON ls.warehouse_id = w.id
        LEFT JOIN cgmf cg ON ls.cgmf_id = cg.id
        WHERE ls.id = ?
    ''', (slip_id,))
    
    if not slip:
        flash('Loading slip not found', 'error')
        return redirect(url_for('warehouse_loading_slips'))
    
    return render_template('print_loading_slip.html', slip=slip[0])

@app.route('/warehouse/print-builty/<string:builty_id>')
@login_required
def warehouse_print_builty(builty_id):
    """Warehouse can print builty"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get builty details with all related information
    builty = db.get_builty_by_id(builty_id)
    
    if not builty:
        flash('Builty not found', 'error')
        return redirect(url_for('warehouse_dashboard'))
    
    # Convert to dict for easier template access
    # Explicit column order from get_builty_by_id query
    builty_dict = {
        'builty_id': builty[0],
        'builty_number': builty[1],
        'rake_code': builty[2],
        'date': builty[3],
        'rake_point_name': builty[4],
        'account_id': builty[5],
        'warehouse_id': builty[6],
        'cgmf_id': builty[7],
        'truck_id': builty[8],
        'loading_point': builty[9],
        'unloading_point': builty[10],
        'goods_name': builty[11],
        'number_of_bags': builty[12],
        'quantity_mt': builty[13],
        'kg_per_bag': builty[14],
        'rate_per_mt': builty[15],
        'total_freight': builty[16],
        'advance': builty[17],
        'to_pay': builty[18],
        'lr_number': builty[19],
        'lr_index': builty[20],
        'created_by_role': builty[21],
        'created_at': builty[22],
        'account_name': builty[23],
        'warehouse_name': builty[24],
        'truck_number': builty[25],
        'driver_name': builty[26],
        'driver_mobile': builty[27],
        'owner_name': builty[28],
        'owner_mobile': builty[29],
        'builty_head': builty[30] if len(builty) > 30 else None,
        'receiver_name': builty[31] if len(builty) > 31 else None,
        'received_quantity': builty[32] if len(builty) > 32 else None,
        'account_address': builty[33] if len(builty) > 33 else None,
        'company_name': builty[34] if len(builty) > 34 else None,
        'company_code': builty[35] if len(builty) > 35 else None,
        'account_type': builty[36] if len(builty) > 36 else None,
        'sub_head': builty[37] if len(builty) > 37 else None,
        'cgmf_society_name': None
    }
    # Lookup CGMF society name if this builty is for a CGMF account
    if builty_dict['cgmf_id']:
        cgmf_row = db.execute_custom_query('SELECT society_name FROM cgmf WHERE id = ?', (builty_dict['cgmf_id'],))
        if cgmf_row:
            builty_dict['cgmf_society_name'] = cgmf_row[0][0]
    # Override account_type to 'Payal' if account name contains 'payal'
    if builty_dict['account_name'] and 'payal' in str(builty_dict['account_name']).lower():
        builty_dict['account_type'] = 'Payal'
    
    return render_template('print_builty.html', builty=builty_dict)

@app.route('/warehouse/create-builty', methods=['GET', 'POST'])
@login_required
def warehouse_create_builty():
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Generate builty number
        builty_number = f"WBLT-{datetime.now().strftime('%Y%m%d')}-{datetime.now().strftime('%H%M%S')}"
        
        loading_slip_id = request.form.get('loading_slip_id')
        warehouse_name = request.form.get('warehouse_name')
        date = request.form.get('date') or request.form.get('builty_date')
        account_warehouse = request.form.get('account_warehouse')
        truck_number = request.form.get('truck_number')
        loading_point = request.form.get('loading_point')
        unloading_point = request.form.get('unloading_point')
        goods_name = request.form.get('goods_name')
        number_of_bags = int(request.form.get('number_of_bags'))
        quantity_wt_mt = float(request.form.get('quantity_wt_mt'))
        freight_details = request.form.get('freight_details')
        lr_number = request.form.get('lr_number')
        
        # Get data from loading slip if provided
        if loading_slip_id:
            loading_slip = db.execute_custom_query('''
                SELECT ls.id, ls.rake_code, ls.slip_number, ls.loading_point_name,
                       ls.destination, ls.quantity_bags, ls.quantity_mt, ls.goods_name,
                       ls.truck_id, ls.truck_driver, ls.truck_owner, 
                       ls.mobile_number_1, ls.mobile_number_2,
                       t.truck_number, t.driver_name, t.driver_mobile, 
                       t.owner_name, t.owner_mobile,
                       ls.account_id, ls.warehouse_id, ls.builty_id
                FROM loading_slips ls
                LEFT JOIN trucks t ON ls.truck_id = t.id
                WHERE ls.id = ?
            ''', (loading_slip_id,))
            
            if loading_slip:
                # Check if this loading slip already has a builty
                if loading_slip[0][20]:  # builty_id is index 20
                    flash('Error: This loading slip already has a builty created!', 'error')
                    return redirect(url_for('warehouse_create_builty'))
                slip = loading_slip[0]
                # Use truck data from JOIN (if available) or from loading_slip fields
                truck_number = slip[13] if slip[13] else truck_number  # t.truck_number
                truck_driver = slip[14] if slip[14] else slip[9]  # t.driver_name or ls.truck_driver
                mobile_number_1 = slip[15] if slip[15] else slip[11]  # t.driver_mobile or ls.mobile_number_1
                truck_owner = slip[16] if slip[16] else slip[10]  # t.owner_name or ls.truck_owner
                mobile_number_2 = slip[17] if slip[17] else slip[12]  # t.owner_mobile or ls.mobile_number_2
                truck_id = slip[8]  # truck_id from loading_slips
                
                # Override form values with loading slip data
                warehouse_name = warehouse_name  # Keep from form as user already selected
                loading_point = slip[3]  # loading_point_name
                goods_name = slip[7]  # goods_name
                number_of_bags = slip[5]  # quantity_bags
                quantity_wt_mt = slip[6]  # quantity_mt
        else:
            # Get truck details from form
            truck_driver = request.form.get('truck_driver')
            truck_owner = request.form.get('truck_owner')
            mobile_number_1 = request.form.get('mobile_number_1')
            mobile_number_2 = request.form.get('mobile_number_2', '')
            
            truck = db.get_truck_by_number(truck_number)
            if not truck:
                truck_id = db.add_truck(truck_number, truck_driver, mobile_number_1, truck_owner, mobile_number_2)
            else:
                truck_id = truck[0]
        
        # Get warehouse ID for stock OUT
        warehouse = db.execute_custom_query('SELECT id FROM warehouses WHERE warehouse_name = ?', (warehouse_name,))
        if not warehouse:
            flash('Invalid warehouse', 'error')
            return redirect(url_for('warehouse_create_builty'))
        warehouse_id = warehouse[0][0]
        
        # Check warehouse balance
        balance = db.get_warehouse_balance_stock(warehouse_id)
        if balance['balance'] < quantity_wt_mt:
            flash(f'Error: Warehouse has only {balance["balance"]:.2f} MT available. Cannot dispatch {quantity_wt_mt:.2f} MT', 'error')
            return redirect(url_for('warehouse_create_builty'))
        
        # Determine destination (account, warehouse, or CGMF)
        accounts = db.get_all_accounts()
        warehouses_list = db.get_all_warehouses()
        account_id = None
        destination_warehouse_id = None
        cgmf_id = None
        
        # Check if it's a CGMF account
        if account_warehouse and account_warehouse.startswith('CGMF:'):
            cgmf_id = account_warehouse.split(':')[1]
        else:
            for account in accounts:
                if account[1] == account_warehouse:
                    account_id = account[0]
                    break
            if account_id is None:
                for wh in warehouses_list:
                    if wh[1] == account_warehouse:
                        destination_warehouse_id = wh[0]
                        break
        
        # Warehouse builties are secondary dispatches; NULL rake_code
        # ensures they do not inflate rake balance/dispatch totals.
        rake_code = None

        # Calculate freight fields
        try:
            total_freight = float(freight_details.replace('₹', '').replace(',', '').strip())
        except:
            total_freight = 0.0
        quantity_wt_mt = float(quantity_wt_mt)
        kg_per_bag = (quantity_wt_mt * 1000) / number_of_bags if number_of_bags > 0 else 50
        rate_per_mt = total_freight / quantity_wt_mt if quantity_wt_mt > 0 else 0
        
        # Get sub_head and receiver details
        sub_head = request.form.get('sub_head', '')
        receiver_name = request.form.get('receiver_name', '')
        received_quantity = request.form.get('received_quantity')
        received_quantity = float(received_quantity) if received_quantity else None
        supply_term = request.form.get('supply_term', 'EX')
        
        # Create builty
        builty_id = db.add_builty(builty_number, rake_code, date, warehouse_name, account_id, destination_warehouse_id,
                                  truck_id, loading_point, unloading_point, goods_name, number_of_bags,
                                  quantity_wt_mt, kg_per_bag, rate_per_mt, total_freight, 0, 0, lr_number,
                                  0, 'warehouse', cgmf_id, sub_head, receiver_name, received_quantity, supply_term)
        
        if builty_id:
            # Link loading slip to builty if provided
            if loading_slip_id:
                db.link_loading_slip_to_builty(loading_slip_id, builty_id)
            
            # Add stock OUT transaction
            stock_id = db.add_warehouse_stock_out(warehouse_id, builty_id, quantity_wt_mt,
                                                  account_id, date, '')
            
            if stock_id:
                flash(f'Builty {builty_number} created and stock OUT recorded successfully!', 'success')
                return redirect(url_for('warehouse_dashboard'))
            else:
                flash('Builty created but error recording stock OUT', 'warning')
                return redirect(url_for('warehouse_dashboard'))
        else:
            flash('Error creating builty', 'error')
    
    # GET request
    warehouses = db.get_all_warehouses()
    accounts = db.get_all_accounts()
    loading_slips = db.get_warehouse_loading_slips()  # Only loading slips created from warehouses
    cgmf_list = db.get_all_cgmf()
    
    return render_template('warehouse/create_builty.html',
                         warehouses=warehouses,
                         accounts=accounts,
                         loading_slips=loading_slips,
                         cgmf_list=cgmf_list)

@app.route('/warehouse/stock-out', methods=['GET', 'POST'])
@login_required
def warehouse_stock_out():
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        warehouse_name = request.form.get('warehouse_name')
        stock_out_date = request.form.get('stock_out_date')
        
        # Builty details (14 fields)
        date = request.form.get('date')
        rake_point_name = request.form.get('rake_point_name')
        account_warehouse = request.form.get('account_warehouse')
        truck_number = request.form.get('truck_number')
        loading_point = request.form.get('loading_point')
        unloading_point = request.form.get('unloading_point')
        truck_driver = request.form.get('truck_driver')
        truck_owner = request.form.get('truck_owner')
        mobile_number_1 = request.form.get('mobile_number_1')
        mobile_number_2 = request.form.get('mobile_number_2', '')
        goods_name = request.form.get('goods_name')
        number_of_bags = int(request.form.get('number_of_bags'))
        quantity_wt_mt = float(request.form.get('quantity_wt_mt'))
        freight_details = request.form.get('freight_details')
        lr_number = request.form.get('lr_number')
        
        # Get warehouse ID
        warehouse = db.execute_custom_query('SELECT id FROM warehouses WHERE warehouse_name = ?', (warehouse_name,))
        if not warehouse:
            flash('Invalid warehouse', 'error')
            return redirect(url_for('warehouse_stock_out'))
        warehouse_id = warehouse[0][0]
        
        # Check if warehouse has enough stock
        balance = db.get_warehouse_balance_stock(warehouse_id)
        if balance['balance'] < quantity_wt_mt:
            flash(f'Error: Warehouse has only {balance["balance"]:.2f} MT available. Cannot dispatch {quantity_wt_mt:.2f} MT', 'error')
            return redirect(url_for('warehouse_stock_out'))
        
        # Determine if account or warehouse
        account_id = None
        destination_warehouse_id = None
        accounts = db.get_all_accounts()
        for account in accounts:
            if account[1] == account_warehouse:
                account_id = account[0]
                break
        
        # Get or create truck
        truck = db.get_truck_by_number(truck_number)
        if not truck:
            truck_id = db.add_truck(truck_number, truck_driver, mobile_number_1, truck_owner, mobile_number_2)
        else:
            truck_id = truck[0]
        
        # Auto-generate builty number for stock OUT
        from datetime import datetime
        builty_number = f"BLTO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Calculate additional fields
        try:
            total_freight = float(freight_details.replace('₹', '').replace(',', '').strip())
        except:
            total_freight = 0.0
        kg_per_bag = (quantity_wt_mt * 1000) / number_of_bags if number_of_bags > 0 else 50
        rate_per_mt = total_freight / quantity_wt_mt if quantity_wt_mt > 0 else 0
        
        # Get rake_code from warehouse stock (use most recent rake)
        rake_code_result = db.execute_custom_query('''
            SELECT b.rake_code
            FROM warehouse_stock ws
            JOIN builty b ON ws.builty_id = b.id
            WHERE ws.warehouse_id = ? AND ws.transaction_type = 'IN'
            ORDER BY ws.created_at DESC
            LIMIT 1
        ''', (warehouse_id,))
        rake_code = rake_code_result[0][0] if rake_code_result else None
        
        # Create builty for stock OUT
        builty_id = db.add_builty(builty_number, rake_code, date, rake_point_name, account_id, destination_warehouse_id,
                                  truck_id, loading_point, unloading_point, goods_name, number_of_bags,
                                  quantity_wt_mt, kg_per_bag, rate_per_mt, total_freight, lr_number,
                                  0, 'warehouse')
        
        if builty_id:
            # Add stock out transaction
            stock_id = db.add_warehouse_stock_out(warehouse_id, builty_id, quantity_wt_mt,
                                                  account_id, stock_out_date, '')
            
            if stock_id:
                flash(f'Stock OUT recorded successfully! Builty: {builty_number}', 'success')
                return redirect(url_for('warehouse_dashboard'))
            else:
                flash('Error recording stock OUT', 'error')
        else:
            flash('Error creating builty for stock OUT', 'error')
    
    warehouses = db.get_all_warehouses()
    accounts = db.get_all_accounts()
    trucks = db.get_all_trucks()
    
    return render_template('warehouse/stock_out.html', 
                         warehouses=warehouses,
                         accounts=accounts,
                         trucks=trucks)

@app.route('/warehouse/balance')
@login_required
def warehouse_balance_all():
    """View balance across all warehouses with filtering support"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get filter parameters
    selected_warehouse = request.args.get('warehouse', 'all')
    date_filter = request.args.get('date_filter', 'all')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Get all warehouses for dropdown
    warehouses = db.get_all_warehouses()
    
    # Build date condition
    date_condition = ""
    date_params = []
    
    from datetime import datetime, timedelta
    today = datetime.now().strftime('%Y-%m-%d')
    
    if date_filter == 'today':
        date_condition = " AND DATE(ws.date) = ?"
        date_params = [today]
    elif date_filter == 'week':
        week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        date_condition = " AND DATE(ws.date) >= ?"
        date_params = [week_ago]
    elif date_filter == 'month':
        month_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        date_condition = " AND DATE(ws.date) >= ?"
        date_params = [month_ago]
    elif date_filter == 'custom' and start_date and end_date:
        date_condition = " AND DATE(ws.date) >= ? AND DATE(ws.date) <= ?"
        date_params = [start_date, end_date]

    # Support both legacy and modern PK naming in related master tables
    def _pk_col(table_name, legacy_pk):
        try:
            cols = db.execute_custom_query(
                f"SELECT name FROM pragma_table_info('{table_name}')"
            ) or []
            names = [c[0] for c in cols if c and len(c) > 0]
            if names:
                return 'id' if 'id' in names else legacy_pk
        except Exception:
            pass
        try:
            cols = db.execute_custom_query(
                f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'"
            ) or []
            names = [c[0] for c in cols if c and len(c) > 0]
            if names:
                return 'id' if 'id' in names else legacy_pk
        except Exception:
            pass
        return legacy_pk

    account_pk = _pk_col('accounts', 'account_id')
    cgmf_pk = _pk_col('cgmf', 'cgmf_id')
    company_pk = _pk_col('companies', 'company_id')
    
    # Get overall stock statistics
    warehouse_balances = []
    account_balances = []
    
    # Build warehouse condition
    if selected_warehouse != 'all':
        try:
            selected_warehouse_id = selected_warehouse
            warehouse_filter = [w for w in warehouses if w[0] == selected_warehouse_id]
            warehouses_to_process = warehouse_filter
        except:
            warehouses_to_process = warehouses
    else:
        warehouses_to_process = warehouses
    
    for warehouse in warehouses_to_process:
        # Get filtered balance
        if date_condition:
            result = db.execute_custom_query(f'''
                SELECT 
                    COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity_mt ELSE 0 END), 0) as stock_in,
                    COALESCE(SUM(CASE WHEN transaction_type = 'OUT' THEN quantity_mt ELSE 0 END), 0) as stock_out
                FROM warehouse_stock ws
                WHERE warehouse_id = ? {date_condition}
            ''', [warehouse[0]] + date_params)
            if result:
                stock_in = result[0][0] or 0
                stock_out = result[0][1] or 0
            else:
                stock_in, stock_out = 0, 0
        else:
            balance = db.get_warehouse_balance_stock(warehouse[0])
            stock_in = balance.get('stock_in', 0)
            stock_out = balance.get('stock_out', 0)
        
        warehouse_balances.append([
            warehouse[0],  # warehouse_id
            warehouse[1],  # warehouse_name
            warehouse[2],  # location
            stock_in,
            stock_out
        ])
    
    # Calculate totals
    total_stock_in = sum(w[3] for w in warehouse_balances)
    total_stock_out = sum(w[4] for w in warehouse_balances)
    total_balance = total_stock_in - total_stock_out
    
    # Get recent transactions for the filtered scope
    if selected_warehouse != 'all':
        wh_cond = "ws.warehouse_id = ?"
        wh_params = [selected_warehouse]
    else:
        wh_cond = "1=1"
        wh_params = []
    
    recent_transactions = db.execute_custom_query(f'''
        SELECT ws.date, ws.transaction_type, ws.quantity_mt, w.warehouse_name,
               COALESCE(b.builty_number, 'Direct Entry') as builty_number,
               ws.notes
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.id
        LEFT JOIN builty b ON ws.builty_id = b.id
        WHERE {wh_cond} {date_condition}
        ORDER BY ws.date DESC, ws.created_at DESC
        LIMIT 20
    ''', wh_params + date_params)
    
    # Get account balances (simplified)
    accounts = db.get_all_accounts()
    for account in accounts:
        account_balances.append([account[1], 0, 0])  # Placeholder
    
    # Get selected warehouse name for display
    selected_warehouse_name = "All Warehouses"
    if selected_warehouse != 'all':
        for wh in warehouses:
            if str(wh[0]) == str(selected_warehouse):
                selected_warehouse_name = wh[1]
                break
    
    return render_template('warehouse/balance.html',
                         warehouse_balances=warehouse_balances,
                         account_balances=account_balances,
                         total_stock_in=total_stock_in,
                         total_stock_out=total_stock_out,
                         total_balance=total_balance,
                         warehouses=warehouses,
                         selected_warehouse=selected_warehouse,
                         selected_warehouse_name=selected_warehouse_name,
                         date_filter=date_filter,
                         start_date=start_date,
                         end_date=end_date,
                         recent_transactions=recent_transactions)

@app.route('/warehouse/balance/<string:warehouse_id>')
@login_required
def warehouse_balance(warehouse_id):
    """View balance for specific warehouse"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    warehouse = db.get_warehouse_by_id(warehouse_id)
    balance = db.get_warehouse_balance_stock(warehouse_id)
    transactions = db.get_warehouse_stock_transactions(warehouse_id)
    
    return render_template('warehouse/balance.html', 
                         warehouse=warehouse,
                         balance=balance,
                         transactions=transactions)

@app.route('/warehouse/balance-summary/<string:warehouse_id>')
@login_required
def warehouse_balance_summary(warehouse_id):
    """Warehouse detail page: account-wise and product-wise stock summary"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))

    warehouse = db.get_warehouse_by_id(warehouse_id)
    if not warehouse:
        flash('Warehouse not found', 'error')
        return redirect(url_for('warehouse_balance_all'))

    date_filter = request.args.get('date_filter', 'all')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    date_condition = ""
    date_params = []
    if date_filter == 'today':
        date_condition = " AND ws.date = ?"
        date_params = [datetime.now().strftime('%Y-%m-%d')]
    elif date_filter == 'week':
        date_condition = " AND ws.date >= ?"
        date_params = [(datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')]
    elif date_filter == 'month':
        date_condition = " AND ws.date >= ?"
        date_params = [(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')]
    elif date_filter == 'custom' and start_date and end_date:
        date_condition = " AND ws.date >= ? AND ws.date <= ?"
        date_params = [start_date, end_date]

    # account_type is a PostgreSQL ENUM — must cast to TEXT before string operations
    _acct_rows = db.execute_custom_query(
        '''SELECT a.account_name,
                  CASE WHEN LOWER(a.account_type::TEXT) IN ('company','payal') THEN 'Payal'
                       WHEN LOWER(a.account_type::TEXT) = 'dealer' THEN 'Dealer'
                       WHEN LOWER(a.account_type::TEXT) = 'retailer' THEN 'Retailer'
                       ELSE a.account_type::TEXT END,
                  COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0)
           FROM warehouse_stock ws
           JOIN accounts a ON ws.account_id = a.id
           WHERE ws.warehouse_id = ? AND ws.account_id IS NOT NULL''' + date_condition + '''
           GROUP BY a.account_name, a.account_type
           HAVING COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
           ORDER BY 3 DESC, 1''',
        [warehouse_id] + date_params
    ) or []

    _cgmf_rows = db.execute_custom_query(
        '''SELECT c.society_name, 'CGMF',
                  COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0)
           FROM warehouse_stock ws
           JOIN cgmf c ON ws.cgmf_id = c.id
           WHERE ws.warehouse_id = ? AND ws.cgmf_id IS NOT NULL''' + date_condition + '''
           GROUP BY c.society_name
           HAVING COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
           ORDER BY 3 DESC, 1''',
        [warehouse_id] + date_params
    ) or []

    _company_rows = db.execute_custom_query(
        '''SELECT co.company_name, 'Company',
                  COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0)
           FROM warehouse_stock ws
           JOIN companies co ON ws.company_id = co.id
           WHERE ws.warehouse_id = ? AND ws.account_id IS NULL AND ws.cgmf_id IS NULL''' + date_condition + '''
           GROUP BY co.company_name
           HAVING COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
           ORDER BY 3 DESC, 1''',
        [warehouse_id] + date_params
    ) or []

    account_stock = _acct_rows + _cgmf_rows + _company_rows

    # Per-account product breakdown (for inline expansion in template)
    _acct_prod = db.execute_custom_query(
        '''SELECT a.account_name, COALESCE(p.product_name, 'N/A'),
                  COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0)
           FROM warehouse_stock ws
           JOIN accounts a ON ws.account_id = a.id
           LEFT JOIN products p ON ws.product_id = p.id
           WHERE ws.warehouse_id = ? AND ws.account_id IS NOT NULL''' + date_condition + '''
           GROUP BY a.account_name, COALESCE(p.product_name, 'N/A')
           HAVING COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
           ORDER BY 1, 3 DESC''',
        [warehouse_id] + date_params
    ) or []

    _cgmf_prod = db.execute_custom_query(
        '''SELECT c.society_name, COALESCE(p.product_name, 'N/A'),
                  COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0)
           FROM warehouse_stock ws
           JOIN cgmf c ON ws.cgmf_id = c.id
           LEFT JOIN products p ON ws.product_id = p.id
           WHERE ws.warehouse_id = ? AND ws.cgmf_id IS NOT NULL''' + date_condition + '''
           GROUP BY c.society_name, COALESCE(p.product_name, 'N/A')
           HAVING COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
           ORDER BY 1, 3 DESC''',
        [warehouse_id] + date_params
    ) or []

    _co_prod = db.execute_custom_query(
        '''SELECT co.company_name, COALESCE(p.product_name, 'N/A'),
                  COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0)
           FROM warehouse_stock ws
           JOIN companies co ON ws.company_id = co.id
           LEFT JOIN products p ON ws.product_id = p.id
           WHERE ws.warehouse_id = ? AND ws.account_id IS NULL AND ws.cgmf_id IS NULL''' + date_condition + '''
           GROUP BY co.company_name, COALESCE(p.product_name, 'N/A')
           HAVING COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
           ORDER BY 1, 3 DESC''',
        [warehouse_id] + date_params
    ) or []

    # Build dict: entity_name -> [(product, qty), ...]
    account_products = {}
    for row in (_acct_prod + _cgmf_prod + _co_prod):
        name = row[0]
        account_products.setdefault(name, []).append((row[1], float(row[2] or 0)))

    product_stock = db.execute_custom_query(
        '''SELECT COALESCE(p.product_name, 'N/A'),
                  COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0)
           FROM warehouse_stock ws
           LEFT JOIN products p ON ws.product_id = p.id
           WHERE ws.warehouse_id = ?''' + date_condition + '''
           GROUP BY COALESCE(p.product_name, 'N/A')
           HAVING COALESCE(SUM(CASE WHEN ws.transaction_type::TEXT = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
           ORDER BY 2 DESC, 1''',
        [warehouse_id] + date_params
    ) or []

    total_account_stock = sum(float(row[2] or 0) for row in account_stock)
    total_product_stock = sum(float(row[1] or 0) for row in product_stock)

    return render_template(
        'warehouse/balance_summary.html',
        warehouse=warehouse,
        account_stock=account_stock,
        account_products=account_products,
        product_stock=product_stock,
        total_account_stock=total_account_stock,
        total_product_stock=total_product_stock,
        date_filter=date_filter,
        start_date=start_date,
        end_date=end_date
    )

@app.route('/warehouse/do-creation', methods=['GET', 'POST'])
@login_required
def warehouse_do_creation():
    """Stock Allotment - Reassign stock from one account to another within warehouse"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        from_account_raw = request.form.get('from_account')
        to_account_raw = request.form.get('to_account')
        quantity = float(request.form.get('quantity', 0))
        warehouse_id = request.form.get('warehouse_id')
        allotment_date = request.form.get('allotment_date')
        notes = request.form.get('notes', '')
        
        # Parse FROM entity (can be account, COMPANY, or CGMF)
        from_account_id = None
        from_cgmf_id = None
        from_company_id = None
        from_display_name = ""
        
        if from_account_raw.startswith('CGMF:'):
            from_cgmf_id = from_account_raw.replace('CGMF:', '')
            cgmf_info = db.execute_custom_query('SELECT society_name FROM cgmf WHERE id = ?', (from_cgmf_id,))
            from_display_name = cgmf_info[0][0] if cgmf_info else f'CGMF #{from_cgmf_id}'
        elif from_account_raw.startswith('COMPANY:'):
            from_company_id = from_account_raw.replace('COMPANY:', '')
            company_info = db.execute_custom_query('SELECT company_name FROM companies WHERE id = ?', (from_company_id,))
            from_display_name = company_info[0][0] if company_info else f'Company #{from_company_id}'
        else:
            from_account_id = from_account_raw
            account_info = db.execute_custom_query('SELECT account_name FROM accounts WHERE id = ?', (from_account_id,))
            from_display_name = account_info[0][0] if account_info else f'Account #{from_account_id}'
        
        # Parse TO entity (can be account, COMPANY, or CGMF)
        to_account_id = None
        to_cgmf_id = None
        to_company_id = None
        to_display_name = ""
        
        if to_account_raw.startswith('CGMF:'):
            to_cgmf_id = to_account_raw.replace('CGMF:', '')
            cgmf_info = db.execute_custom_query('SELECT society_name FROM cgmf WHERE id = ?', (to_cgmf_id,))
            to_display_name = cgmf_info[0][0] if cgmf_info else f'CGMF #{to_cgmf_id}'
        elif to_account_raw.startswith('COMPANY:'):
            to_company_id = to_account_raw.replace('COMPANY:', '')
            company_info = db.execute_custom_query('SELECT company_name FROM companies WHERE id = ?', (to_company_id,))
            to_display_name = company_info[0][0] if company_info else f'Company #{to_company_id}'
        else:
            to_account_id = to_account_raw
            account_info = db.execute_custom_query('SELECT account_name FROM accounts WHERE id = ?', (to_account_id,))
            to_display_name = account_info[0][0] if account_info else f'Account #{to_account_id}'
        
        # Validate
        if not all([from_account_raw, to_account_raw, quantity > 0, warehouse_id]):
            flash('Please fill all required fields', 'error')
            return redirect(url_for('warehouse_do_creation'))
        
        if from_account_raw == to_account_raw:
            flash('FROM and TO accounts cannot be the same', 'error')
            return redirect(url_for('warehouse_do_creation'))
        
        try:
            # Check available stock for FROM entity in this warehouse
            if from_account_id:
                available = db.execute_custom_query('''
                    SELECT COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity_mt ELSE -quantity_mt END), 0)
                    FROM warehouse_stock WHERE warehouse_id = ? AND account_id = ?
                ''', (warehouse_id, from_account_id))
            elif from_cgmf_id:
                available = db.execute_custom_query('''
                    SELECT COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity_mt ELSE -quantity_mt END), 0)
                    FROM warehouse_stock WHERE warehouse_id = ? AND cgmf_id = ?
                ''', (warehouse_id, from_cgmf_id))
            elif from_company_id:
                available = db.execute_custom_query('''
                    SELECT COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity_mt ELSE -quantity_mt END), 0)
                    FROM warehouse_stock WHERE warehouse_id = ? AND company_id = ? AND account_id IS NULL AND cgmf_id IS NULL
                ''', (warehouse_id, from_company_id))
            else:
                available = [[0]]
            
            available_qty = available[0][0] if available and available[0][0] else 0
            
            if quantity > available_qty:
                flash(f'Insufficient stock! Available: {available_qty:.2f} MT', 'error')
                return redirect(url_for('warehouse_do_creation'))
            
            # Record stock OUT from source entity
            if from_account_id:
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, account_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, 'OUT', ?, ?, 'allotment')
                ''', (warehouse_id, from_account_id, quantity, allotment_date, f'Allotment to {to_display_name}: {notes}'))
            elif from_cgmf_id:
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, cgmf_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, 'OUT', ?, ?, 'allotment')
                ''', (warehouse_id, from_cgmf_id, quantity, allotment_date, f'Allotment to {to_display_name}: {notes}'))
            elif from_company_id:
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, company_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, 'OUT', ?, ?, 'allotment')
                ''', (warehouse_id, from_company_id, quantity, allotment_date, f'Allotment to {to_display_name}: {notes}'))
            
            # Record stock IN to destination entity
            if to_account_id:
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, account_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, 'IN', ?, ?, 'allotment')
                ''', (warehouse_id, to_account_id, quantity, allotment_date, f'Allotment from {from_display_name}: {notes}'))
            elif to_cgmf_id:
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, cgmf_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, 'IN', ?, ?, 'allotment')
                ''', (warehouse_id, to_cgmf_id, quantity, allotment_date, f'Allotment from {from_display_name}: {notes}'))
            elif to_company_id:
                db.execute_custom_query('''
                    INSERT INTO warehouse_stock (warehouse_id, company_id, quantity_mt, transaction_type, date, remark, source_type)
                    VALUES (?, ?, ?, 'IN', ?, ?, 'allotment')
                ''', (warehouse_id, to_company_id, quantity, allotment_date, f'Allotment from {from_display_name}: {notes}'))
            
            flash(f'Stock allotment successful! {quantity:.2f} MT transferred from {from_display_name} to {to_display_name}.', 'success')
            return redirect(url_for('warehouse_do_creation'))
            
        except Exception as e:
            print(f"Error in stock allotment: {e}")
            flash('Error processing stock allotment', 'error')
            return redirect(url_for('warehouse_do_creation'))
    
    # Support both legacy schema (*_id PKs) and newer schema (id PKs).
    def _pk_col(table_name, legacy_pk):
        try:
            # SQLite / libsql path
            cols = db.execute_custom_query(
                f"SELECT name FROM pragma_table_info('{table_name}')"
            ) or []
            names = [c[0] for c in cols if c and len(c) > 0]
            if names:
                return 'id' if 'id' in names else legacy_pk
        except Exception:
            pass

        try:
            # PostgreSQL path
            cols = db.execute_custom_query(
                f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'"
            ) or []
            names = [c[0] for c in cols if c and len(c) > 0]
            if names:
                return 'id' if 'id' in names else legacy_pk
        except Exception:
            pass

        return legacy_pk

    account_pk = _pk_col('accounts', 'account_id')
    cgmf_pk = _pk_col('cgmf', 'cgmf_id')
    company_pk = _pk_col('companies', 'company_id')
    product_pk = _pk_col('products', 'product_id')
    warehouse_pk = _pk_col('warehouses', 'warehouse_id')

    # GET request - fetch entities with positive stock (product-wise) for allotment source table/dropdown.
    account_stock = db.execute_custom_query(f'''
        SELECT w.{warehouse_pk}, w.warehouse_name, a.{account_pk}, a.account_name,
               CASE
                   WHEN LOWER(COALESCE(a.account_type, '')) IN ('company', 'payal') THEN 'Payal'
                   WHEN LOWER(COALESCE(a.account_type, '')) = 'dealer' THEN 'Dealer'
                   WHEN LOWER(COALESCE(a.account_type, '')) = 'retailer' THEN 'Retailer'
                   ELSE COALESCE(a.account_type, 'Account')
               END AS account_type,
               COALESCE(p.product_name, 'N/A') AS product_name,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.{warehouse_pk}
        JOIN accounts a ON ws.account_id = a.{account_pk}
        LEFT JOIN products p ON ws.product_id = p.{product_pk}
        WHERE ws.account_id IS NOT NULL
        GROUP BY w.{warehouse_pk}, w.warehouse_name, a.{account_pk}, a.account_name,
                 CASE
                     WHEN LOWER(COALESCE(a.account_type, '')) IN ('company', 'payal') THEN 'Payal'
                     WHEN LOWER(COALESCE(a.account_type, '')) = 'dealer' THEN 'Dealer'
                     WHEN LOWER(COALESCE(a.account_type, '')) = 'retailer' THEN 'Retailer'
                     ELSE COALESCE(a.account_type, 'Account')
                 END,
                 COALESCE(p.product_name, 'N/A')
        HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ORDER BY w.warehouse_name, a.account_name, COALESCE(p.product_name, 'N/A')
    ''') or []
    
    cgmf_stock = db.execute_custom_query(f'''
        SELECT w.{warehouse_pk}, w.warehouse_name, c.{cgmf_pk}, c.society_name, 'CGMF' as type,
               COALESCE(p.product_name, 'N/A') AS product_name,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.{warehouse_pk}
        JOIN cgmf c ON ws.cgmf_id = c.{cgmf_pk}
        LEFT JOIN products p ON ws.product_id = p.{product_pk}
        WHERE ws.cgmf_id IS NOT NULL
        GROUP BY w.{warehouse_pk}, w.warehouse_name, c.{cgmf_pk}, c.society_name, COALESCE(p.product_name, 'N/A')
        HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ORDER BY w.warehouse_name, c.society_name, COALESCE(p.product_name, 'N/A')
    ''') or []
    
    company_stock = db.execute_custom_query(f'''
        SELECT w.{warehouse_pk}, w.warehouse_name, c.{company_pk}, c.company_name, 'Company' as type,
               COALESCE(p.product_name, 'N/A') AS product_name,
               COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) as balance
        FROM warehouse_stock ws
        JOIN warehouses w ON ws.warehouse_id = w.{warehouse_pk}
        JOIN companies c ON ws.company_id = c.{company_pk}
        LEFT JOIN products p ON ws.product_id = p.{product_pk}
        WHERE ws.company_id IS NOT NULL AND ws.account_id IS NULL AND ws.cgmf_id IS NULL
        GROUP BY w.{warehouse_pk}, w.warehouse_name, c.{company_pk}, c.company_name, COALESCE(p.product_name, 'N/A')
        HAVING COALESCE(SUM(CASE WHEN ws.transaction_type = 'IN' THEN ws.quantity_mt ELSE -ws.quantity_mt END), 0) > 0
        ORDER BY w.warehouse_name, c.company_name, COALESCE(p.product_name, 'N/A')
    ''') or []
    
    # Combine all stock lists
    all_stock = account_stock + cgmf_stock + company_stock
    
    # Get all allotment history
    allotment_history = db.execute_custom_query('''
        SELECT ws.date, a.account_name, a.account_type,
               ws.quantity_mt, ws.transaction_type, w.warehouse_name, ws.remark
        FROM warehouse_stock ws
        JOIN accounts a ON ws.account_id = a.id
        JOIN warehouses w ON ws.warehouse_id = w.id
        WHERE ws.source_type = 'allotment'
        ORDER BY ws.created_at DESC
        LIMIT 50
    ''') or []
    
    accounts = db.get_all_accounts()
    warehouses = db.get_all_warehouses()
    companies = db.get_all_companies()
    cgmf_list = db.execute_custom_query('SELECT cgmf_id, society_name, district FROM cgmf ORDER BY society_name') or []
    
    return render_template('warehouse/do_creation.html',
                         account_stock=all_stock,
                         allotment_history=allotment_history,
                         accounts=accounts,
                         warehouses=warehouses,
                         companies=companies,
                         cgmf_list=cgmf_list)

@app.route('/warehouse/all-builties')
@login_required
def warehouse_all_builties():
    """View all builties created by warehouse"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get all builties created by warehouse role OR warehouse-specific builties
    # Warehouse builties have builty numbers starting with "WBLT-" or created_by_role='Warehouse'
    builties = db.execute_custom_query('''
        SELECT b.id, b.builty_number, b.rake_code, b.date, b.rake_point_name,
               b.account_id, b.warehouse_id, b.cgmf_id, b.truck_id,
               b.loading_point, b.unloading_point, b.goods_name, b.number_of_bags,
               b.quantity_mt, b.kg_per_bag, b.rate_per_mt, b.total_freight,
               b.advance, b.to_pay, b.lr_number, b.lr_index, b.created_by_role,
               b.sub_head, b.receiver_name, b.received_quantity, b.created_at,
               a.account_name, w.warehouse_name, t.truck_number
        FROM builty b
        LEFT JOIN accounts a ON b.account_id = a.id
        LEFT JOIN warehouses w ON b.warehouse_id = w.id
        LEFT JOIN trucks t ON b.truck_id = t.id
        WHERE b.created_by_role = 'warehouse' OR b.builty_number LIKE 'WBLT-%%'
        ORDER BY b.created_at DESC
    ''') or []
    
    return render_template('warehouse/all_builties.html', builties=builties)

@app.route('/warehouse/view-ebills')
@login_required
def warehouse_view_ebills():
    """Warehouse can view e-bills for builties they created"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    ebills = db.get_ebills_by_builty_creator('warehouse')
    return render_template('warehouse/view_ebills.html', ebills=ebills)

@app.route('/warehouse/download-bill/<filename>')
@login_required
def warehouse_download_bill(filename):
    """Warehouse can download bill PDFs for their builties"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('warehouse_view_ebills'))

@app.route('/warehouse/download-eway-bill/<filename>')
@login_required
def warehouse_download_eway_bill(filename):
    """Warehouse can download eway bill PDFs for their builties"""
    if current_user.role != 'Warehouse':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'eway_bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('warehouse_view_ebills'))

# ========== ACCOUNTANT Dashboard & Routes ==========

@app.route('/accountant/dashboard')
@login_required
def accountant_dashboard():
    if current_user.role != 'Accountant':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    # Get all e-bills and recent ones
    all_ebills = db.get_all_ebills()
    recent_ebills = all_ebills[:10] if all_ebills else []
    
    # Get builties without e-bills
    pending_builties = db.get_builties_without_ebills()
    
    # Calculate statistics
    from datetime import datetime, date
    
    # Total e-bills count
    ebill_count = len(all_ebills) if all_ebills else 0
    
    # E-bills with eway bill PDFs uploaded
    eway_bill_count = sum(1 for ebill in all_ebills if ebill[4]) if all_ebills else 0
    
    # Total builties available
    all_builties = db.get_all_builties()
    builty_count = len(all_builties) if all_builties else 0
    
    # Today's e-bills
    today = date.today().strftime('%Y-%m-%d')
    today_ebill_count = sum(1 for ebill in all_ebills if ebill[3] and ebill[3].startswith(today)) if all_ebills else 0
    
    # Total accounts
    all_accounts = db.get_all_accounts()
    account_count = len(all_accounts) if all_accounts else 0
    
    # Pending uploads (e-bills without PDF)
    pending_uploads = ebill_count - eway_bill_count
    
    # This month's e-bills
    current_month = date.today().strftime('%Y-%m')
    month_ebills = sum(1 for ebill in all_ebills if ebill[3] and ebill[3].startswith(current_month)) if all_ebills else 0
    
    return render_template('accountant/dashboard.html', 
                         ebills=all_ebills,
                         recent_ebills=recent_ebills,
                         pending_builties=pending_builties,
                         ebill_count=ebill_count,
                         eway_bill_count=eway_bill_count,
                         builty_count=builty_count,
                         today_ebill_count=today_ebill_count,
                         account_count=account_count,
                         pending_uploads=pending_uploads,
                         month_ebills=month_ebills)

@app.route('/accountant/create-ebill', methods=['GET', 'POST'])
@login_required
def accountant_create_ebill():
    if current_user.role != 'Accountant':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        builty_id = request.form.get('builty_id')
        ebill_number = request.form.get('ebill_number')
        amount = 0.0  # Amount field removed from form
        generated_date = request.form.get('ebill_date')  # Fixed: was 'generated_date'
        
        # Handle file upload for bill PDF
        bill_pdf = None
        if 'bill_pdf' in request.files:
            file = request.files['bill_pdf']
            if file and file.filename and file.filename.endswith('.pdf'):
                # Create uploads directory if it doesn't exist
                import os
                upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'bills')
                os.makedirs(upload_folder, exist_ok=True)
                
                # Generate unique filename
                from datetime import datetime
                filename = f"BILL_{ebill_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                filepath = os.path.join(upload_folder, filename)
                
                # Save the file
                file.save(filepath)
                bill_pdf = filename
        
        # Handle file upload for eway bill PDF
        eway_bill_pdf = None
        if 'eway_bill_pdf' in request.files:
            file = request.files['eway_bill_pdf']
            if file and file.filename and file.filename.endswith('.pdf'):
                # Create uploads directory if it doesn't exist
                import os
                upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'eway_bills')
                os.makedirs(upload_folder, exist_ok=True)
                
                # Generate unique filename
                from datetime import datetime
                filename = f"EWAY_{ebill_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                filepath = os.path.join(upload_folder, filename)
                
                # Save the file
                file.save(filepath)
                eway_bill_pdf = filename
        
        ebill_id = db.add_ebill(builty_id, ebill_number, amount, generated_date, bill_pdf, eway_bill_pdf)
        
        if ebill_id:
            flash(f'E-Bill {ebill_number} created successfully!', 'success')
            return redirect(url_for('accountant_dashboard'))
        else:
            flash('Error creating e-bill. E-Bill number may already exist or date is missing.', 'error')
    
    builties = db.get_builties_without_ebills()
    
    return render_template('accountant/create_ebill.html', builties=builties)

@app.route('/accountant/ebills')
@login_required
def accountant_all_ebills():
    if current_user.role != 'Accountant':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    ebills = db.get_all_ebills()
    
    return render_template('accountant/all_ebills.html', ebills=ebills)

@app.route('/accountant/download-eway-bill/<filename>')
@login_required
def download_eway_bill(filename):
    """Download eway bill PDF"""
    if current_user.role != 'Accountant':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'eway_bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('accountant_all_ebills'))

@app.route('/accountant/download-bill/<filename>')
@login_required
def download_bill(filename):
    """Download bill PDF"""
    if current_user.role != 'Accountant':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    filename = os.path.basename(filename)  # Prevent path traversal
    upload_folder = os.path.join(os.path.dirname(__file__), 'uploads', 'bills')
    filepath = os.path.join(upload_folder, filename)
    
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        flash('File not found', 'error')
        return redirect(url_for('accountant_all_ebills'))


@app.route('/admin/download-database-backup')
@login_required
def download_database():
    """Admin-only endpoint to download database backup"""
    if current_user.role != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('index'))
    
    db_path = os.path.join(os.path.dirname(__file__), 'fims.db')
    if os.path.exists(db_path):
        # Create backup filename with timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(db_path, as_attachment=True, download_name=f'fims_backup_{timestamp}.db')
    else:
        flash('Database file not found', 'error')
        return redirect(url_for('admin_dashboard'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
